"""Feature A: Data export to Excel — orders, finance summary, settlements."""

from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select, extract, func, or_
from sqlalchemy.orm import selectinload
from datetime import date

from app.core.datetime_helpers import today_cn, to_cn
from decimal import Decimal
from typing import Optional
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app.core.deps import DBSession, CurrentUser
from app.models.order import Order, OrderStatus, Channel
from app.models.settlement import OwnerSettlement
from app.models.expense import Expense
from app.models.owner import Owner
from app.models.room import Room
from app.models.order_room import OrderRoom
from app.services.order_pricing import (
    effective_owner_revenue_for_room,
    order_room_commission,
    safe_decimal,
)

router = APIRouter(prefix="/export", tags=["export"])

# ── Helpers ──────────────────────────────────────────────────────────────────

HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1677FF", end_color="1677FF", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

CHANNEL_LABELS = {
    # 现役渠道（2026-05-23 拍板 + trial_stay）
    "ctrip": "携程", "meituan_hotel": "美团酒店", "meituan_homestay": "美团民宿",
    "douyin": "抖音", "qunar": "去哪儿", "zhixing": "智行", "tongcheng": "同程",
    "self_acquired": "自来客", "offline": "线下渠道", "self_used": "自住",
    "trial_stay": "试住",
    # 历史保留值（前端隐藏，仅供老订单展示）
    "meituan": "美团（旧）", "tujia": "途家", "private": "私域",
    "walk_in": "散客", "direct": "自助",
}

STATUS_LABELS = {
    # 2026-06-05 流程调整：确认收款挪到退房后。
    # pending_checkout=已退房待收款，pending_payment 重定义为后期「待完成」。
    "pending_confirm": "待确认", "pending_payment": "待完成",
    "paid_pending_room": "已付待排房", "roomed_pending_checkin": "待入住",
    "checked_in": "入住中", "pending_checkout": "已退房待收款",
    "completed": "已完成", "cancelled": "已取消",
    "rescheduled": "已改期", "abnormal": "异常",
}


def _style_header(ws, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _to_streaming(wb, filename: str) -> StreamingResponse:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    # HTTP 头只能 latin-1；中文文件名走 RFC 6266 filename*，并给 ASCII 兜底名。
    from urllib.parse import quote

    ascii_fallback = filename.encode("ascii", "ignore").decode("ascii") or "export.xlsx"
    disposition = (
        f"attachment; filename=\"{ascii_fallback}\"; "
        f"filename*=UTF-8''{quote(filename)}"
    )
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": disposition},
    )


# ── Export Orders ────────────────────────────────────────────────────────────

@router.get("/orders")
async def export_orders(
    db: DBSession,
    current_user: CurrentUser,
    status: Optional[str] = Query(default=None),
    check_in_from: Optional[date] = Query(default=None),
    check_in_to: Optional[date] = Query(default=None),
):
    """Export filtered orders as Excel."""
    if current_user["role"] not in ("admin", "operator", "finance"):
        from fastapi import HTTPException
        raise HTTPException(status_code=403, detail="无权导出")

    q = select(Order).options(selectinload(Order.rooms)).where(Order.is_deleted == False)
    if status:
        q = q.where(Order.order_status == status)
    if check_in_from:
        q = q.where(Order.check_in_date >= check_in_from)
    if check_in_to:
        q = q.where(Order.check_in_date <= check_in_to)
    q = q.order_by(Order.created_at.desc())

    result = await db.execute(q)
    orders = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "订单列表"

    headers = ["订单号", "渠道", "客人姓名", "手机号", "房间", "入住日期", "退房日期",
               "晚数", "实收金额", "佣金", "净收入", "预计收入(业主到手)", "状态", "创建时间"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for o in orders:
        ws.append([
            o.order_id,
            CHANNEL_LABELS.get(o.channel.value, o.channel.value),
            o.guest_name,
            o.guest_phone,
            # 多房订单读 order_rooms（room_id 顶层字段已废弃，多房单恒 NULL → 曾显示"待排房"）
            "、".join(o.room_ids) or "待排房",
            str(o.check_in_date),
            str(o.check_out_date),
            o.nights,
            float(o.actual_price or 0),
            float(o.platform_commission),
            float(o.net_revenue),
            float(o.expected_revenue) if o.expected_revenue is not None else "",
            STATUS_LABELS.get(o.order_status.value, o.order_status.value),
            to_cn(o.created_at).strftime("%Y-%m-%d %H:%M") if o.created_at else "",
        ])

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    today_str = today_cn().strftime("%Y%m%d")
    return _to_streaming(wb, f"orders_{today_str}.xlsx")


# ── Export Finance Report ────────────────────────────────────────────────────

@router.get("/finance")
async def export_finance(
    db: DBSession,
    current_user: CurrentUser,
    year: int = Query(default_factory=lambda: today_cn().year),
    month: int = Query(default_factory=lambda: today_cn().month),
    start_date: Optional[date] = Query(default=None, description="起始日期 YYYY-MM-DD"),
    end_date: Optional[date] = Query(default=None, description="结束日期 YYYY-MM-DD"),
    floor: Optional[int] = Query(default=None, description="楼层"),
    owner_id: Optional[str] = Query(default=None, description="业主ID"),
    room_id: Optional[str] = Query(default=None, description="房间ID"),
):
    """导出财务报表 Excel。日期范围二选一（start_date+end_date 优先，向后兼容 year+month）。
    收入按订单离店日、支出按发生日落区间，与月度汇总/结算口径一致。"""
    if current_user["role"] not in ("admin", "finance"):
        from fastapi import HTTPException
        raise HTTPException(status_code=403, detail="无权导出财务数据")

    use_range = bool(start_date and end_date)
    if use_range and end_date < start_date:
        from fastapi import HTTPException
        raise HTTPException(status_code=400, detail="end_date 必须 >= start_date")

    # 收入：必须按 OrderRoom（房段）离店日和金额导出。顶层 Order 对多房、换房单
    # 只有整单汇总值，直接导出会跨楼层串账并漏掉平台补贴。
    order_date_filter = (
        [OrderRoom.check_out_date >= start_date, OrderRoom.check_out_date <= end_date]
        if use_range
        else [extract("year", OrderRoom.check_out_date) == year,
              extract("month", OrderRoom.check_out_date) == month]
    )
    income_q = (
        select(OrderRoom, Order, Room)
        .join(Order, Order.order_id == OrderRoom.order_id)
        .join(Room, Room.room_id == OrderRoom.room_id)
        .where(
            Order.is_deleted == False,
            Order.order_status.not_in([OrderStatus.cancelled]),
            *order_date_filter,
        )
    )
    if floor is not None:
        income_q = income_q.where(Room.floor == floor)
    if owner_id:
        income_q = income_q.where(Room.owner_id == owner_id)
    if room_id:
        income_q = income_q.where(Room.room_id == room_id)
    income_rows = (await db.execute(
        income_q.order_by(OrderRoom.check_out_date, Room.room_name, Order.order_id)
    )).all()

    order_ids = {orow.order_id for orow, _, _ in income_rows}
    room_counts: dict[str, int] = {}
    order_room_names: dict[str, list[str]] = {}
    if order_ids:
        room_counts = dict((await db.execute(
            select(OrderRoom.order_id, func.count())
            .where(OrderRoom.order_id.in_(order_ids))
            .group_by(OrderRoom.order_id)
        )).all())
        all_room_names = (await db.execute(
            select(OrderRoom.order_id, Room.room_name)
            .join(Room, Room.room_id == OrderRoom.room_id)
            .where(OrderRoom.order_id.in_(order_ids))
            .order_by(OrderRoom.order_id, OrderRoom.position)
        )).all()
        for oid, room_name in all_room_names:
            order_room_names.setdefault(oid, []).append(room_name)

    # 支出：区间按发生日 between，否则按 year/month
    expense_date_filter = (
        [Expense.expense_date >= start_date, Expense.expense_date <= end_date]
        if use_range
        else [extract("year", Expense.expense_date) == year,
              extract("month", Expense.expense_date) == month]
    )
    expense_q = select(Expense).outerjoin(Room, Room.room_id == Expense.room_id).where(
            *expense_date_filter,
            Expense.is_deleted == False,
        )
    if floor is not None:
        # 整层导出无法可靠分摊 room_id 为空的业主级支出，留给正式结算包展示。
        expense_q = expense_q.where(Room.floor == floor)
    if owner_id:
        expense_q = expense_q.where(or_(Expense.owner_id == owner_id, Room.owner_id == owner_id))
    if room_id:
        expense_q = expense_q.where(Expense.room_id == room_id)
    expenses_result = await db.execute(expense_q.order_by(Expense.expense_date))
    expenses = expenses_result.scalars().all()

    wb = openpyxl.Workbook()

    # Sheet 1: Revenue summary
    ws1 = wb.active
    ws1.title = "收入明细"
    h1 = ["订单号", "房段ID", "渠道", "客人", "楼层", "房间", "整单房间", "入住", "退房", "晚数",
          "房费", "平台佣金", "平台补贴", "分尾调整", "结算净收入"]
    ws1.append(h1)
    _style_header(ws1, len(h1))

    # 结算表头采用账务惯例「每个房号先舍到分，再相加」。逐订单行各自舍入时，
    # 同一房号可能产生 1~N 分尾差；把尾差明确落在该房最后一行，保证导出行合计
    # 与按房号页/正式结算严格一致，而不是留下看不见的 0.01~0.03 差额。
    pricing: dict[str, tuple[Decimal, Decimal, Decimal, Decimal]] = {}
    indexes_by_room: dict[str, list[str]] = {}
    raw_net_by_room: dict[str, Decimal] = {}
    for orow, o, _room in income_rows:
        actual = Decimal(orow.actual_price or 0)
        per_room_net = effective_owner_revenue_for_room(
            orow.ota_owner_revenue, orow.actual_price, o.metadata_,
            room_counts.get(o.order_id),
        )
        commission = order_room_commission(
            orow.actual_price, per_room_net, o.platform_commission_rate)
        subsidy = (
            safe_decimal((o.metadata_ or {}).get("ota_subsidy"))
            if per_room_net is None else Decimal("0")
        )
        raw_net = actual - commission + subsidy
        pricing[orow.order_room_id] = (actual, commission, subsidy, raw_net)
        indexes_by_room.setdefault(orow.room_id, []).append(orow.order_room_id)
        raw_net_by_room[orow.room_id] = raw_net_by_room.get(orow.room_id, Decimal("0")) + raw_net

    tail_adjustment: dict[str, Decimal] = {}
    cent = Decimal("0.01")
    for rid, row_ids in indexes_by_room.items():
        desired = raw_net_by_room[rid].quantize(cent)
        displayed = sum((pricing[row_id][3].quantize(cent) for row_id in row_ids), Decimal("0"))
        tail_adjustment[row_ids[-1]] = desired - displayed

    total_actual = Decimal("0")
    total_commission = Decimal("0")
    total_subsidy = Decimal("0")
    total_adjustment = Decimal("0")
    total_net = Decimal("0")
    for orow, o, room in income_rows:
        actual, commission_raw, subsidy_raw, net_raw = pricing[orow.order_room_id]
        commission = commission_raw.quantize(cent)
        subsidy = subsidy_raw.quantize(cent)
        adjustment = tail_adjustment.get(orow.order_room_id, Decimal("0"))
        net = net_raw.quantize(cent) + adjustment
        ws1.append([
            o.order_id,
            orow.order_room_id,
            CHANNEL_LABELS.get(o.channel.value, o.channel.value),
            (o.guest_name[:1] + "**") if o.guest_name else "",
            room.floor if room.floor is not None else "",
            room.room_name,
            "、".join(order_room_names.get(o.order_id, [])),
            str(orow.check_in_date),
            str(orow.check_out_date),
            max((orow.check_out_date - orow.check_in_date).days, 0),
            float(actual),
            float(commission),
            float(subsidy),
            float(adjustment),
            float(net),
        ])
        total_actual += actual
        total_commission += commission
        total_subsidy += subsidy
        total_adjustment += adjustment
        total_net += net

    # Summary row
    ws1.append([])
    ws1.append(["合计", "", "", "", "", "", "", "", "", "", float(total_actual),
                float(total_commission), float(total_subsidy), float(total_adjustment), float(total_net)])

    for col in ws1.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    # Sheet 2: Expenses
    ws2 = wb.create_sheet("支出明细")
    h2 = ["日期", "类别", "描述", "房间", "金额"]
    ws2.append(h2)
    _style_header(ws2, len(h2))

    total_expense = Decimal("0")
    for e in expenses:
        ws2.append([
            str(e.expense_date) if e.expense_date else "",
            e.category.value if hasattr(e.category, "value") else str(e.category),
            e.description,
            e.room_id or "",
            float(e.amount),
        ])
        total_expense += e.amount

    ws2.append([])
    ws2.append(["合计", "", "", "", float(total_expense)])

    for col in ws2.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    fname = (
        f"finance_{start_date.strftime('%Y%m%d')}-{end_date.strftime('%Y%m%d')}.xlsx"
        if use_range
        else f"finance_{year}{str(month).zfill(2)}.xlsx"
    )
    return _to_streaming(wb, fname)


# ── Export Settlements ───────────────────────────────────────────────────────

@router.get("/settlements")
async def export_settlements(
    db: DBSession,
    current_user: CurrentUser,
    billing_month: Optional[str] = Query(default=None),
):
    """Export owner settlements as Excel."""
    if current_user["role"] not in ("admin", "finance"):
        from fastapi import HTTPException
        raise HTTPException(status_code=403, detail="无权导出结算数据")

    q = select(OwnerSettlement).options(selectinload(OwnerSettlement.items))
    if billing_month:
        q = q.where(OwnerSettlement.billing_month == billing_month)
    q = q.order_by(OwnerSettlement.billing_month.desc())

    result = await db.execute(q)
    settlements = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "房东结算"

    # 房东分成/实际结算导出到厘（与页面一致）；净收入/扣除支出到分。
    from app.services.owner_settlement import settlement_precise_amounts
    headers = ["结算ID", "房东ID", "结算月", "净收入", "房东分成", "扣除支出", "实际结算", "状态", "备注"]
    ws.append(headers)
    _style_header(ws, len(headers))

    status_labels = {"pending": "待确认", "confirmed": "已确认", "paid": "已打款", "disputed": "有争议"}

    for s in settlements:
        owner_amount_p, actual_p = settlement_precise_amounts(s.items)
        # 无明细的旧版结算 precise 为 None → 回退到分（存库值），不显示 0
        owner_amount_out = float(owner_amount_p if owner_amount_p is not None else s.owner_amount)
        actual_out = float(actual_p if actual_p is not None else s.actual_owner_amount)
        ws.append([
            s.settlement_id,
            s.owner_id,
            s.billing_month,
            float(s.total_net_revenue),
            owner_amount_out,
            float(s.deducted_expenses),
            actual_out,
            status_labels.get(s.status.value if hasattr(s.status, "value") else s.status, str(s.status)),
            s.notes or "",
        ])
        # 显式数字格式：分成/实际结算(E/G 列)到厘，净收入/扣除支出(D/F 列)到分。
        row = ws.max_row
        ws.cell(row=row, column=4).number_format = "0.00"
        ws.cell(row=row, column=5).number_format = "0.000"
        ws.cell(row=row, column=6).number_format = "0.00"
        ws.cell(row=row, column=7).number_format = "0.000"

    # 表尾说明：与页面小字一致，避免业主拿导出对不上到账金额。
    ws.append([])
    ws.append(["说明：房东分成 / 实际结算精确至厘（元），实际打款按分四舍五入结算。"])

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    month_str = billing_month or today_cn().strftime("%Y-%m")
    return _to_streaming(wb, f"settlements_{month_str}.xlsx")


# ── Export Single Owner Settlement Statement (每行「导出明细」) ────────────────
#
# 单个业主结算单 → 『业主分成明细 + 垫付费用明细』两段式 Excel，可直接发给业主。
# 规格（王总 2026-07-19，精简版）：模式写「业主X%」；垫付明细只填金额、数量/备注留空
# 手填；无高亮、无活公式，数值直接算好。关键不变量：主表每行
#   业主分成 + Σ各支出列(负) + 运营补贴 == 净利润，且 Σ各支出列 == 业主总支出。

# 主表支出列顺序（有数据源的固定列 + 「其他」兜底，保证 Σ列 == 业主总支出，不漏账）
_STATEMENT_EXPENSE_COLUMNS = [
    "日耗", "保洁", "续住保洁", "洗涤", "物业费", "水费", "公摊水费",
    "电费", "新布草过水费", "维修费", "燃气费", "采购费", "其他",
]
# 覆盖 v2#7 全部在用类目（对齐王总《业主分成明细》），未命中才落「其他」。
# 改支出类目枚举时必须同步这里，否则新类目会全堆进「其他」。见 expense.py ExpenseCategory。
_CATEGORY_TO_COLUMN = {
    "daily_supplies": "日耗",
    "laundry": "洗涤",
    "property_fee": "物业费",
    "water": "水费",
    "cold_water": "水费", "hot_water": "水费",  # v2#7 前旧数据
    "public_utilities": "公摊水费",
    "electricity": "电费",
    "new_linen_prewash": "新布草过水费",
    "maintenance": "维修费",
    "gas": "燃气费",
    "supplies": "采购费",
    "kitchen_cleaning": "保洁",          # 厨房保洁归保洁
    "property_guidance_fee": "物业费",   # 物业引导费归物业
    # broadband / utilities(旧水电统称) / other → 「其他」兜底
}
_STATEMENT_COL_WIDTH = {
    "序号": 5, "模式": 9, "户型": 24, "小区名字": 18, "房号": 14,
    "房费": 11, "业主分成": 11, "新布草过水费": 12, "运营补贴": 10,
    "业主总支出": 12, "净利润": 12, "收款账号": 34,
}


def _statement_column_for(category: str, description: str) -> str:
    """把某笔业主支出归到主表哪一列。cleaning 按描述含「续住」拆保洁/续住保洁；
    BookingType 无 continuation 值，只能靠描述兜底（无「续住」字样默认落保洁，仍对账）。
    未识别类目一律落「其他」，保证 Σ各列 == 业主总支出。"""
    if category == "cleaning":
        return "续住保洁" if "续住" in (description or "") else "保洁"
    return _CATEGORY_TO_COLUMN.get(category, "其他")


def _stmt_round(x) -> Decimal:
    return Decimal(str(x or 0)).quantize(Decimal("0.01"))


def _build_statement_wb(settlement, owner, rooms_map: dict, desc_map: dict):
    """rooms_map: room_id -> (room_id, room_name, room_type, community_name)
    desc_map:  expense_id -> description（拆分保洁/续住保洁用）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "分成明细"

    header = (
        ["序号", "模式", "户型", "小区名字", "房号", "房费", "业主分成"]
        + _STATEMENT_EXPENSE_COLUMNS
        + ["运营补贴", "业主总支出", "净利润", "收款账号"]
    )
    ncol = len(header)

    # 标题行 + 表头行
    ws.append([f"{settlement.billing_month} 业主分成明细"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    tcell = ws.cell(row=1, column=1)
    tcell.font = Font(bold=True, size=13)
    tcell.alignment = Alignment(horizontal="center", vertical="center")

    ws.append(header)
    for c in range(1, ncol + 1):
        cell = ws.cell(row=2, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # 收款账号（户名/账号/开户行）
    payee = ""
    if owner:
        parts = [f"户名：{owner.name}"]
        if owner.bank_account:
            parts.append(f"账号：{owner.bank_account}")
        if owner.bank_name:
            parts.append(f"开户行：{owner.bank_name}")
        payee = "\n".join(parts)

    # 整个业主折成一行（王总 7-19：跟样表一样，一个业主一行合计）。
    # 逐 item 汇总金额 + 收集元数据（户型/小区去重、房号取范围、模式取综合比例）。
    import re

    totals = {k: Decimal("0") for k in _STATEMENT_EXPENSE_COLUMNS}
    category_totals: dict = {}  # 垫付明细用：column -> owner_amount 累计（正数）
    tot_fee = tot_share = tot_subsidy = tot_expense = tot_net = Decimal("0")
    room_types: list = []
    communities: list = []
    room_nums: list = []
    room_names_nonnum: list = []

    for i in settlement.items:
        room = rooms_map.get(i.room_id)
        rt = (room[2] if room else None) or ""
        cm = (room[3] if room else None) or ""
        if rt and rt not in room_types:
            room_types.append(rt)
        if cm and cm not in communities:
            communities.append(cm)
        if i.room_id:  # 业主级支出行（room_id 空）不进房号
            nm = (room[1] if room else None) or i.room_id or ""
            m = re.search(r"\d+", nm)
            if m:
                room_nums.append(int(m.group()))
            elif nm:
                room_names_nonnum.append(nm)

        tot_fee += _stmt_round(i.net_revenue)
        tot_expense += _stmt_round(i.owner_expenses)
        tot_net += _stmt_round(i.owner_net_amount)
        tot_share += _stmt_round(_stmt_round(i.owner_net_amount) + _stmt_round(i.owner_expenses))

        for e in (i.cost_share_breakdown or []):
            if e.get("category") == "平台补贴" or e.get("type") == "platform_subsidy":
                tot_subsidy += _stmt_round(e.get("amount"))
                continue
            oa = _stmt_round(e.get("owner_amount"))
            col = _statement_column_for(
                e.get("category", ""), desc_map.get(e.get("expense_id"), "")
            )
            totals[col] += oa
            if oa:
                category_totals[col] = category_totals.get(col, Decimal("0")) + oa

    # 模式：所有房同比例 → 「业主X%」；否则按综合比例（分成/房费）反推
    ratios = {Decimal(str(i.share_ratio_snapshot or 0)) for i in settlement.items if i.room_id}
    if len(ratios) == 1:
        mode = f"业主{int(next(iter(ratios)) * 100)}%"
    elif tot_fee > 0:
        mode = f"业主{int((tot_share / tot_fee * 100).to_integral_value())}%"
    else:
        mode = ""

    # 房号范围：数字房号取 min-max；有非数字（业主级 label 已排除）则补「等」
    if room_nums:
        lo, hi = min(room_nums), max(room_nums)
        room_cell = f"{lo}-{hi}" if hi > lo else str(lo)
        if room_names_nonnum:
            room_cell += "等"
    else:
        room_cell = "、".join(room_names_nonnum)

    row = [1, mode, "、".join(room_types), "、".join(communities), room_cell,
           float(tot_fee), float(tot_share)]
    for k in _STATEMENT_EXPENSE_COLUMNS:
        v = totals[k]
        row.append(float(-v) if v else "")   # 支出列显示负数；0 留空
    row.append(float(tot_subsidy) if tot_subsidy else "")
    row.append(float(-tot_expense))
    row.append(float(tot_net))
    row.append(payee)
    ws.append(row)

    # 数据行：垂直居中，长文本列自动换行，行高给足（收款账号多行）
    data_row = ws.max_row
    for c in range(1, ncol + 1):
        wrap = header[c - 1] in ("户型", "小区名字", "收款账号")
        ws.cell(row=data_row, column=c).alignment = Alignment(vertical="center", wrap_text=wrap)
    ws.row_dimensions[data_row].height = 46

    # 列宽（中文列加宽，避免截断）
    for idx, name in enumerate(header, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = (
            _STATEMENT_COL_WIDTH.get(name, 9)
        )

    # ── 垫付费用明细：接在主表下方（同一 sheet，对齐样表上下堆叠）──
    r0 = data_row + 2  # 主表下空一行再起
    tcell = ws.cell(row=r0, column=1, value="垫付费用明细")
    ws.merge_cells(start_row=r0, start_column=1, end_row=r0, end_column=5)
    tcell.font = Font(bold=True, size=13)
    tcell.alignment = Alignment(horizontal="center", vertical="center")

    detail_header = ["序号", "费用名称", "费用数量", "金额", "备注"]
    for c, name in enumerate(detail_header, start=1):
        cell = ws.cell(row=r0 + 1, column=c, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    rr = r0 + 2
    dseq = 0
    for k in _STATEMENT_EXPENSE_COLUMNS:  # 按主表列顺序，只列有金额的类目
        amt = category_totals.get(k)
        if not amt:
            continue
        dseq += 1
        ws.cell(row=rr, column=1, value=dseq)
        ws.cell(row=rr, column=2, value=k)
        # 费用数量(C)/备注(E) 留空手填（王总口径）；金额(D)
        ws.cell(row=rr, column=4, value=float(amt))
        rr += 1
    if dseq == 0:
        ws.cell(row=rr, column=1, value=1)
        ws.cell(row=rr, column=2, value="（本月无垫付费用）")

    return wb


async def _settlement_export_context(db, settlement_id: str):
    """加载结算快照及展示元数据；所有单结算导出共用。"""
    from fastapi import HTTPException

    settlement = (await db.execute(
        select(OwnerSettlement)
        .options(selectinload(OwnerSettlement.items))
        .where(OwnerSettlement.settlement_id == settlement_id)
    )).scalar_one_or_none()
    if not settlement:
        raise HTTPException(status_code=404, detail="结算记录不存在")
    owner = (await db.execute(
        select(Owner).where(Owner.owner_id == settlement.owner_id)
    )).scalar_one_or_none()
    room_ids = [i.room_id for i in settlement.items if i.room_id]
    rooms_map: dict = {}
    if room_ids:
        rr = await db.execute(
            select(Room.room_id, Room.room_name, Room.room_type, Room.community_name)
            .where(Room.room_id.in_(room_ids))
        )
        rooms_map = {row[0]: row for row in rr.all()}
    expense_ids = {
        e.get("expense_id")
        for i in settlement.items
        for e in (i.cost_share_breakdown or [])
        if e.get("expense_id")
    }
    desc_map: dict = {}
    if expense_ids:
        er = await db.execute(
            select(Expense.expense_id, Expense.description)
            .where(Expense.expense_id.in_(expense_ids))
        )
        desc_map = {row[0]: row[1] for row in er.all()}
    return settlement, owner, rooms_map, desc_map


async def _settlement_income_rows(db, settlement):
    """按当前源数据重建收入行，并逐房核对结算快照。

    结算单是不可变账务快照；源订单若在生成后被修改，拒绝导出看似精确但实际
    对不上的明细，要求先重新生成待确认结算单。
    """
    from calendar import monthrange
    from fastapi import HTTPException

    year, month = (int(x) for x in settlement.billing_month.split("-"))
    start = date(year, month, 1)
    end = date(year, month, monthrange(year, month)[1])
    snapshot = {
        i.room_id: Decimal(i.net_revenue).quantize(Decimal("0.01"))
        for i in settlement.items if i.room_id
    }
    room_ids = list(snapshot)
    q = (
        select(OrderRoom, Order, Room)
        .join(Order, Order.order_id == OrderRoom.order_id)
        .join(Room, Room.room_id == OrderRoom.room_id)
        .where(
            OrderRoom.room_id.in_(room_ids),
            OrderRoom.check_out_date >= start,
            OrderRoom.check_out_date <= end,
            Order.is_deleted == False,
            Order.order_status != OrderStatus.cancelled,
        )
        .order_by(Room.room_name, OrderRoom.check_out_date, Order.order_id)
    )
    source_rows = (await db.execute(q)).all() if room_ids else []
    order_ids = {o.order_id for _, o, _ in source_rows}
    room_counts: dict[str, int] = {}
    if order_ids:
        room_counts = dict((await db.execute(
            select(OrderRoom.order_id, func.count())
            .where(OrderRoom.order_id.in_(order_ids))
            .group_by(OrderRoom.order_id)
        )).all())

    rows: list[dict] = []
    raw_by_room = {rid: Decimal("0") for rid in room_ids}
    row_indexes_by_room: dict[str, list[int]] = {rid: [] for rid in room_ids}
    order_ids_by_room: dict[str, set[str]] = {rid: set() for rid in room_ids}
    for orow, order, room in source_rows:
        actual = Decimal(orow.actual_price or 0)
        per_room_net = effective_owner_revenue_for_room(
            orow.ota_owner_revenue, orow.actual_price, order.metadata_,
            room_counts.get(order.order_id),
        )
        commission = order_room_commission(
            orow.actual_price, per_room_net, order.platform_commission_rate)
        subsidy = (
            safe_decimal((order.metadata_ or {}).get("ota_subsidy"))
            if per_room_net is None else Decimal("0")
        )
        raw_net = actual - commission + subsidy
        raw_by_room[orow.room_id] += raw_net
        order_ids_by_room[orow.room_id].add(order.order_id)
        row_indexes_by_room[orow.room_id].append(len(rows))
        rows.append({
            "order_id": order.order_id,
            "order_room_id": orow.order_room_id,
            "channel": CHANNEL_LABELS.get(order.channel.value, order.channel.value),
            "guest": (order.guest_name[:1] + "**") if order.guest_name else "",
            "room_id": room.room_id,
            "room_name": room.room_name,
            "check_in": orow.check_in_date,
            "check_out": orow.check_out_date,
            "nights": max((orow.check_out_date - orow.check_in_date).days, 0),
            "actual": actual.quantize(Decimal("0.01")),
            "commission": commission.quantize(Decimal("0.01")),
            "subsidy": subsidy.quantize(Decimal("0.01")),
            "net": raw_net.quantize(Decimal("0.01")),
        })

    check_rows = []
    for rid, snap in snapshot.items():
        current = raw_by_room[rid].quantize(Decimal("0.01"))
        if current != snap:
            raise HTTPException(
                status_code=409,
                detail=(f"结算数据已变化（房间 {rid}：快照 {snap}，当前 {current}），"
                        "请先重新生成待确认结算单再导出。"),
            )
        # 单行先舍入可能产生分尾差，将尾差落在该房最后一行，使明细逐行相加严格等于快照。
        indexes = row_indexes_by_room[rid]
        if indexes:
            line_sum = sum((rows[i]["net"] for i in indexes), Decimal("0"))
            rows[indexes[-1]]["net"] += snap - line_sum
        check_rows.append({
            "room_id": rid,
            "room_name": "",
            "order_count": len(order_ids_by_room[rid]),
            "snapshot": snap,
            "current": current,
            "diff": current - snap,
        })
    return rows, check_rows


def _append_income_detail_sheets(wb, income_rows: list[dict], check_rows: list[dict], rooms_map: dict):
    ws = wb.create_sheet("收入明细")
    headers = ["订单号", "房段ID", "渠道", "客人", "房间ID", "房号", "入住", "退房", "晚数",
               "房费", "平台佣金", "平台补贴", "结算净收入"]
    ws.append(headers)
    _style_header(ws, len(headers))
    for r in income_rows:
        ws.append([
            r["order_id"], r["order_room_id"], r["channel"], r["guest"], r["room_id"],
            r["room_name"], str(r["check_in"]), str(r["check_out"]), r["nights"],
            float(r["actual"]), float(r["commission"]), float(r["subsidy"]), float(r["net"]),
        ])
    ws.append([])
    ws.append(["合计", "", "", "", "", "", "", "", "",
               float(sum((r["actual"] for r in income_rows), Decimal("0"))),
               float(sum((r["commission"] for r in income_rows), Decimal("0"))),
               float(sum((r["subsidy"] for r in income_rows), Decimal("0"))),
               float(sum((r["net"] for r in income_rows), Decimal("0")))])

    ck = wb.create_sheet("房号核对")
    check_headers = ["房间ID", "房号", "订单数", "结算快照净收入", "当前明细净收入", "差额"]
    ck.append(check_headers)
    _style_header(ck, len(check_headers))
    for r in check_rows:
        room = rooms_map.get(r["room_id"])
        ck.append([r["room_id"], room[1] if room else r["room_id"], r["order_count"],
                   float(r["snapshot"]), float(r["current"]), float(r["diff"])])
    ck.append(["合计", "", sum(r["order_count"] for r in check_rows),
               float(sum((r["snapshot"] for r in check_rows), Decimal("0"))),
               float(sum((r["current"] for r in check_rows), Decimal("0"))),
               float(sum((r["diff"] for r in check_rows), Decimal("0")))])

    for sheet in (ws, ck):
        for col in sheet.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            sheet.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)


@router.get("/settlements/{settlement_id}/statement")
async def export_settlement_statement(
    settlement_id: str,
    db: DBSession,
    current_user: CurrentUser,
):
    """单个业主结算 → 『业主分成明细 + 垫付费用明细』Excel（可直接发给业主）。"""
    from fastapi import HTTPException

    if current_user["role"] not in ("admin", "finance"):
        raise HTTPException(status_code=403, detail="无权导出结算明细")

    settlement, owner, rooms_map, desc_map = await _settlement_export_context(
        db, settlement_id)

    wb = _build_statement_wb(settlement, owner, rooms_map, desc_map)
    owner_name = (owner.name if owner else None) or settlement.owner_id
    return _to_streaming(wb, f"业主分成明细_{owner_name}_{settlement.billing_month}.xlsx")


@router.get("/settlements/{settlement_id}/income-detail")
async def export_settlement_income_detail(
    settlement_id: str,
    db: DBSession,
    current_user: CurrentUser,
):
    """导出逐订单房段收入，并保证合计与结算快照逐房一致。"""
    from fastapi import HTTPException

    if current_user["role"] not in ("admin", "finance"):
        raise HTTPException(status_code=403, detail="无权导出结算明细")
    settlement, owner, rooms_map, _ = await _settlement_export_context(db, settlement_id)
    income_rows, check_rows = await _settlement_income_rows(db, settlement)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _append_income_detail_sheets(wb, income_rows, check_rows, rooms_map)
    owner_name = (owner.name if owner else None) or settlement.owner_id
    return _to_streaming(wb, f"到账收入明细_{owner_name}_{settlement.billing_month}.xlsx")


@router.get("/settlements/{settlement_id}/package")
async def export_settlement_package(
    settlement_id: str,
    db: DBSession,
    current_user: CurrentUser,
):
    """一键导出分成表、逐单收入和房号核对三张表。"""
    from fastapi import HTTPException

    if current_user["role"] not in ("admin", "finance"):
        raise HTTPException(status_code=403, detail="无权导出结算明细")
    settlement, owner, rooms_map, desc_map = await _settlement_export_context(db, settlement_id)
    income_rows, check_rows = await _settlement_income_rows(db, settlement)
    wb = _build_statement_wb(settlement, owner, rooms_map, desc_map)
    _append_income_detail_sheets(wb, income_rows, check_rows, rooms_map)
    owner_name = (owner.name if owner else None) or settlement.owner_id
    return _to_streaming(wb, f"完整结算包_{owner_name}_{settlement.billing_month}.xlsx")
