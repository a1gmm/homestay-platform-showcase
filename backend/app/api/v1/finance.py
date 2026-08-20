from fastapi import APIRouter, BackgroundTasks, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func, case
from sqlalchemy.orm import aliased
from typing import Optional
from decimal import Decimal, InvalidOperation
from datetime import date, datetime, timezone
import io
import re
import uuid

from app.core.deps import DBSession, CurrentUser, RedisClient
from app.core.rate_limit import enforce_rate_limit
from app.models.payment import Payment
from app.models.refund import Refund, RefundReason
from app.models.expense import Expense, ExpenseCategory, ExpensePayer, EXPENSE_CATEGORY_LABELS
from app.models.order import Order, OrderStatus, PaymentStatus, is_owner_self_order
from app.models.order_room import OrderRoom
from app.models.room import Room
from app.models.owner import Owner
from app.core.datetime_helpers import today_cn
from app.schemas.finance import (
    PaymentCreate, PaymentUpdate, PaymentOut, RefundCreate, RefundOut,
    ExpenseCreate, ExpenseOut, ByRoomSummaryItem, RoomRevenueSegment,
    ExpenseImportResult,
    RenewalCleaningChargeCreate,
    ServiceFeeConfigOut, ServiceFeeConfigUpdate,
)
from app.services.audit import log_action_tx
from app.services import service_fee_ledger
from app.services.service_fees import get_or_create_config_row, get_service_fees
from app.services.feishu_lead_alert import send_cleaning_renewal_alert
from app.services.payment_service import (
    sum_house_fee_paid, sum_house_fee_refunded, sum_deposit_paid,
)

router = APIRouter(tags=["finance"])


async def _reread_automatic_fee_after_owner_lock(
    db, expense: Expense
) -> Expense | None:
    """Serialize automatic-fee deletion, then refresh its active row.

    Manual expenses retain their existing delete behavior. Automatic service
    fees share the owner-scoped ledger lock with checkout, reconciliation, and
    settlement confirmation; the Expense row is only trusted after that lock.
    """
    if not expense.is_service_fee:
        return expense
    if not expense.owner_id:
        raise HTTPException(
            status_code=409,
            detail="自动服务费缺少业主归属，无法安全作废；请先核查费用数据",
        )

    await service_fee_ledger.lock_owner_service_fee_ledger(db, expense.owner_id)
    return (
        await db.execute(
            select(Expense)
            .where(
                Expense.expense_id == expense.expense_id,
                Expense.is_deleted == False,
            )
            .with_for_update()
            .execution_options(populate_existing=True)
        )
    ).scalar_one_or_none()


# ─── Payments ────────────────────────────────────────────────────────────────

async def _recompute_order_payment_status(db, order_id: str) -> None:
    """重算订单 payment_status —— 收款/退款所有增删改的单一真相源（批2）。

    退款优先：只要还有有效退款(排押金退还口径、排软删)就落 partial_refund/full_refund，
    否则按已收房费判 unpaid/partial/paid。此前该 helper 只认收款侧，改收款备注/删收款
    会把退款单的 *_refund 状态误刷回 paid/partial（评审 F/G）；create/delete refund 的
    状态判定也曾各写一份、与配额校验口径分叉（评审 C/D）。现全部收敛到此。
    列表/聚合统一过滤 is_deleted=True 的收款/退款记录。
    """
    order_result = await db.execute(
        select(Order)
        .where(Order.order_id == order_id, Order.is_deleted == False)
        .with_for_update()
    )
    order = order_result.scalar_one_or_none()
    if not order:
        return
    # 房费口径：排押金(#48)、排软删；退款排押金退还口径 deposit_return。
    total_paid = await sum_house_fee_paid(db, order_id)
    total_refunded = await sum_house_fee_refunded(db, order_id)
    if total_refunded > 0:
        # 有有效退款 → 退款态。full 仅当有应收价且已退≥应收。
        if order.actual_price is not None and total_refunded >= order.actual_price:
            order.payment_status = PaymentStatus.full_refund
        else:
            order.payment_status = PaymentStatus.partial_refund
        return
    # 无有效退款 → 按已收判。无应收价(遗留单)只能粗判：有收款即 partial，否则 unpaid。
    if order.actual_price is None:
        order.payment_status = (
            PaymentStatus.partial if total_paid > 0 else PaymentStatus.unpaid
        )
        return
    if total_paid <= 0:
        order.payment_status = PaymentStatus.unpaid
    elif total_paid >= order.actual_price:
        order.payment_status = PaymentStatus.paid
    else:
        order.payment_status = PaymentStatus.partial


@router.post("/payments", response_model=PaymentOut, status_code=201)
async def create_payment(
    body: PaymentCreate,
    db: DBSession,
    current_user: CurrentUser,
    redis: RedisClient,
):
    if current_user["role"] not in ("admin", "operator"):
        raise HTTPException(status_code=403, detail="无权录入收款")
    await enforce_rate_limit(
        redis, f"payment_create:{current_user['user_id']}", limit=30, window_seconds=60,
        detail="录入过于频繁，请稍后再试",
    )

    # Lock the order row first. Concurrent payments on the same order will serialise here,
    # preventing two transactions from both seeing an outdated total_paid and each accepting
    # a payment that, combined, would exceed actual_price.
    order_result = await db.execute(
        select(Order)
        .where(Order.order_id == body.order_id, Order.is_deleted == False)
        .with_for_update()
    )
    order = order_result.scalar_one_or_none()
    if not order:
        raise HTTPException(status_code=404, detail="订单不存在")

    # Reject payments that would push total_paid past actual_price. If the order has no
    # actual_price yet (legacy/partial data), fall back to old behaviour.
    # 押金不占用房费额度:押金收款(is_deposit)不参与房费上限校验 (#48)。
    if order.actual_price is not None and not body.is_deposit:
        prev_total = await sum_house_fee_paid(db, body.order_id)
        if prev_total + body.amount > order.actual_price:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"收款金额超过订单应收。已收房费 {prev_total}，本次 {body.amount}，"
                    f"应收 {order.actual_price}"
                ),
            )

    # 押金收款上限（批2 item2）：累计押金收款不得超过订单登记押金额 order.deposit。
    # order.deposit=0 视为未登记押金额，不设约束（避免误伤未记押金额却先收押金的流程）。
    if body.is_deposit and order.deposit and order.deposit > 0:
        prev_deposit = await sum_deposit_paid(db, body.order_id)
        if prev_deposit + body.amount > order.deposit:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"押金收款超过订单押金额。已收押金 {prev_deposit}，本次 {body.amount}，"
                    f"应收押金 {order.deposit}"
                ),
            )

    payment = Payment(
        payment_id="PAY-" + uuid.uuid4().hex[:12].upper(),
        created_by=current_user["user_id"],
        **body.model_dump(),
    )
    db.add(payment)
    await db.flush()

    # 重算 payment_status（单一真相源，refund-aware）：对已有退款的订单补收款也不会
    # 把退款态误刷成 paid。
    await _recompute_order_payment_status(db, body.order_id)

    await log_action_tx(db, current_user["user_id"], "payment.create", "payment", payment.payment_id)
    await db.commit()
    await db.refresh(payment)
    return payment


_FINANCE_ROLES = ("admin", "operator", "finance")


@router.get("/payments", response_model=list[PaymentOut])
async def list_payments(
    db: DBSession,
    current_user: CurrentUser,
    order_id: Optional[str] = Query(default=None),
):
    if current_user["role"] not in _FINANCE_ROLES:
        raise HTTPException(status_code=403, detail="无权查看收款记录")
    q = select(Payment).join(Order, Payment.order_id == Order.order_id)
    q = q.where(Order.is_deleted == False, Payment.is_deleted == False)
    if order_id:
        q = q.where(Payment.order_id == order_id)
    q = q.order_by(Payment.paid_at.desc())
    result = await db.execute(q)
    return result.scalars().all()


@router.patch("/payments/{payment_id}", response_model=PaymentOut)
async def update_payment(
    payment_id: str,
    body: PaymentUpdate,
    db: DBSession,
    current_user: CurrentUser,
):
    if current_user["role"] not in ("admin", "operator"):
        raise HTTPException(status_code=403, detail="仅管理员或运营可修改收款记录")

    result = await db.execute(
        select(Payment).where(Payment.payment_id == payment_id, Payment.is_deleted == False)
    )
    payment = result.scalar_one_or_none()
    if not payment:
        raise HTTPException(status_code=404, detail="收款记录不存在")

    updates = body.model_dump(exclude_none=True)
    if not updates:
        raise HTTPException(status_code=400, detail="未提供任何修改字段")

    # 校验新金额不会让订单累计实收超过应收，也不会低于已退（批2 item1/item2）
    if "amount" in updates:
        order_result = await db.execute(
            select(Order)
            .where(Order.order_id == payment.order_id, Order.is_deleted == False)
            .with_for_update()
        )
        order = order_result.scalar_one_or_none()
        if order and not payment.is_deposit:
            # 房费收款：上限≤应收，下限≥已退（改后实收若低于已退，账面变负）。
            others_sum = await sum_house_fee_paid(db, payment.order_id, exclude_payment_id=payment_id)
            new_house_fee = others_sum + updates["amount"]
            if order.actual_price is not None and new_house_fee > order.actual_price:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"修改后累计收款超过订单应收。其他已收 {others_sum}，"
                        f"本次拟改为 {updates['amount']}，应收 {order.actual_price}"
                    ),
                )
            refunded = await sum_house_fee_refunded(db, payment.order_id)
            if new_house_fee < refunded:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"修改后实收房费 {new_house_fee} 低于已退金额 {refunded}，"
                        "会使账面变负。请先冲正相关退款。"
                    ),
                )
        elif order and payment.is_deposit and order.deposit and order.deposit > 0:
            # 押金收款：累计不得超过订单押金额（批2 item2）。
            others_deposit = await sum_deposit_paid(db, payment.order_id, exclude_payment_id=payment_id)
            if others_deposit + updates["amount"] > order.deposit:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"修改后押金收款超过订单押金额。其他已收押金 {others_deposit}，"
                        f"本次拟改为 {updates['amount']}，应收押金 {order.deposit}"
                    ),
                )

    before = {
        "amount": str(payment.amount),
        "method": payment.method.value if payment.method else None,
        "notes": payment.notes,
    }
    for field, value in updates.items():
        setattr(payment, field, value)
    await db.flush()
    await _recompute_order_payment_status(db, payment.order_id)
    after = {
        "amount": str(payment.amount),
        "method": payment.method.value if payment.method else None,
        "notes": payment.notes,
    }
    await log_action_tx(
        db, current_user["user_id"], "payment.update", "payment", payment_id,
        before_data=before, after_data=after,
    )
    await db.commit()
    await db.refresh(payment)
    return payment


@router.delete("/payments/{payment_id}", status_code=204)
async def delete_payment(
    payment_id: str,
    db: DBSession,
    current_user: CurrentUser,
):
    if current_user["role"] not in ("admin", "operator"):
        raise HTTPException(status_code=403, detail="仅管理员或运营可删除收款记录")

    result = await db.execute(
        select(Payment).where(Payment.payment_id == payment_id, Payment.is_deleted == False)
    )
    payment = result.scalar_one_or_none()
    if not payment:
        raise HTTPException(status_code=404, detail="收款记录不存在")

    # 锁订单行，与并发退款/收款串行化——否则删收款的 item1 下限校验会与并发退款建单
    # 各看旧值双双通过，留下「实收<已退」的负账（口径同 create_refund 的行锁）。
    await db.execute(
        select(Order)
        .where(Order.order_id == payment.order_id, Order.is_deleted == False)
        .with_for_update()
    )

    # 对称下限校验（批2 item1）：删掉这笔房费收款后，剩余实收房费不得低于已退金额，
    # 否则账面变负。押金收款不占房费实收，无需此校验。
    if not payment.is_deposit:
        remaining_house_fee = await sum_house_fee_paid(
            db, payment.order_id, exclude_payment_id=payment_id
        )
        refunded = await sum_house_fee_refunded(db, payment.order_id)
        if remaining_house_fee < refunded:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"删除后实收房费 {remaining_house_fee} 低于已退金额 {refunded}，"
                    "会使账面变负。请先冲正相关退款。"
                ),
            )

    before = {
        "amount": str(payment.amount),
        "method": payment.method.value if payment.method else None,
        "order_id": payment.order_id,
    }
    payment.is_deleted = True
    payment.deleted_at = datetime.now(timezone.utc)
    payment.deleted_by = current_user["user_id"]
    await db.flush()
    await _recompute_order_payment_status(db, payment.order_id)
    await log_action_tx(
        db, current_user["user_id"], "payment.delete", "payment", payment_id,
        before_data=before,
    )
    await db.commit()


# ─── Refunds ──────────────────────────────────────────────────────────────────

@router.post("/refunds", response_model=RefundOut, status_code=201)
async def create_refund(body: RefundCreate, db: DBSession, current_user: CurrentUser):
    if current_user["role"] not in ("admin", "operator"):
        raise HTTPException(status_code=403, detail="无权录入退款")

    # 锁住订单行，避免并发退款各自看到旧的 total_refunded 双双通过。
    order_result = await db.execute(
        select(Order)
        .where(Order.order_id == body.order_id, Order.is_deleted == False)
        .with_for_update()
    )
    order = order_result.scalar_one_or_none()
    if not order:
        raise HTTPException(status_code=404, detail="订单不存在")

    # 押金退还有自己的工作流（order.deposit_returned，走退房结算），不在这里录。
    # 放行的话它会挤占房费退款额度（下方 prev_refunded 是房费口径），账本错位。
    if body.reason == RefundReason.deposit_return:
        raise HTTPException(
            status_code=400,
            detail="押金退还请走退房流程的押金结算，不在退款记录中录入。",
        )

    # 累计退款不得超过订单应收。previously 没校验，账面会变负。
    # 排除历史遗留的 deposit_return 行（押金口径）+ 已软删退款（口径同 sum_house_fee_refunded）。
    prev_refunded = await sum_house_fee_refunded(db, body.order_id)
    if order.actual_price is not None and prev_refunded + body.amount > order.actual_price:
        raise HTTPException(
            status_code=400,
            detail=(
                f"退款金额超过订单应收。已退 {prev_refunded}，本次 {body.amount}，"
                f"应收 {order.actual_price}"
            ),
        )

    # 更硬的上限：不能退出比实际收到的还多的钱。押金走 deposit_status 工作流
    # 结算，不计入房费实收；软删的收款不算（口径同 sum_house_fee_paid）。
    received = await sum_house_fee_paid(db, body.order_id)
    if prev_refunded + body.amount > received:
        raise HTTPException(
            status_code=400,
            detail=(
                f"退款金额超过实收房费。实收 {received}，已退 {prev_refunded}，"
                f"本次 {body.amount}。请先核对收款记录。"
            ),
        )

    refund = Refund(
        refund_id="RFD-" + uuid.uuid4().hex[:12].upper(),
        created_by=current_user["user_id"],
        **body.model_dump(),
    )
    db.add(refund)
    await db.flush()

    # 重算 payment_status（单一真相源，refund-aware）：口径与配额校验一致，排 deposit_return。
    await _recompute_order_payment_status(db, body.order_id)

    await log_action_tx(db, current_user["user_id"], "refund.create", "refund", refund.refund_id)
    await db.commit()
    await db.refresh(refund)
    return refund


@router.get("/refunds", response_model=list[RefundOut])
async def list_refunds(db: DBSession, current_user: CurrentUser, order_id: Optional[str] = Query(default=None)):
    if current_user["role"] not in _FINANCE_ROLES:
        raise HTTPException(status_code=403, detail="无权查看退款记录")
    q = select(Refund).join(Order, Refund.order_id == Order.order_id)
    q = q.where(Order.is_deleted == False, Refund.is_deleted == False)
    if order_id:
        q = q.where(Refund.order_id == order_id)
    result = await db.execute(q.order_by(Refund.refunded_at.desc()))
    return result.scalars().all()


@router.delete("/refunds/{refund_id}", status_code=204)
async def delete_refund(
    refund_id: str,
    db: DBSession,
    current_user: CurrentUser,
):
    """软删/冲正退款（批2 item6）：误录的退款只作废、不物理删（保留审计追溯）。
    删后从列表/累计/额度剔除，并重算订单 payment_status。"""
    if current_user["role"] not in ("admin", "operator"):
        raise HTTPException(status_code=403, detail="仅管理员或运营可冲正退款记录")

    result = await db.execute(
        select(Refund).where(Refund.refund_id == refund_id, Refund.is_deleted == False)
    )
    refund = result.scalar_one_or_none()
    if not refund:
        raise HTTPException(status_code=404, detail="退款记录不存在")

    before = {
        "amount": str(refund.amount),
        "reason": refund.reason.value if refund.reason else None,
        "order_id": refund.order_id,
    }
    refund.is_deleted = True
    refund.deleted_at = datetime.now(timezone.utc)
    refund.deleted_by = current_user["user_id"]
    await db.flush()

    # 重算 payment_status（单一真相源，refund-aware；内部对订单行加锁与并发串行化）。
    await _recompute_order_payment_status(db, refund.order_id)

    await log_action_tx(
        db, current_user["user_id"], "refund.delete", "refund", refund_id,
        before_data=before,
    )
    await db.commit()


# ─── Expenses ────────────────────────────────────────────────────────────────

async def _resolve_expense_scope(db, data: dict, *, floor: int | None) -> None:
    """把「整层 / 某业主」录入范围就地解析成 owner_id。

    - floor 有值：取该层所有托管房（有业主）的去重业主。恰 1 个 → owner_id 设该业主、
      room_id 置空（业主级支出，一条不拆）；0 个 → 拒绝；>1 个 → 拒绝并提示改用按业主。
    - 只给 owner_id（无 room_id、无 floor）：校验业主存在。
    - 给了 room_id：单房支出，行为不变，不动。
    data 会被就地修改（写入 owner_id / 清空 room_id）。
    """
    if floor is not None:
        if data.get("room_id"):
            raise HTTPException(status_code=422, detail="整层录入不能同时指定房间")
        owners = (await db.execute(
            select(Room.owner_id)
            .where(
                Room.floor == floor,
                Room.owner_id.isnot(None),
                Room.is_deleted == False,
            )
            .distinct()
        )).scalars().all()
        owners = [o for o in owners if o]
        if not owners:
            raise HTTPException(status_code=422, detail=f"{floor} 层没有托管房间，无法整层记账")
        if len(owners) > 1:
            raise HTTPException(status_code=422, detail=f"{floor} 层有多个业主，请改用「按业主」分别录入")
        data["owner_id"] = owners[0]
        data["room_id"] = None
        return

    if data.get("owner_id") and not data.get("room_id"):
        exists = (await db.execute(
            select(Owner.owner_id).where(Owner.owner_id == data["owner_id"])
        )).scalar_one_or_none()
        if exists is None:
            raise HTTPException(status_code=422, detail="业主不存在")


@router.post("/expenses", response_model=ExpenseOut, status_code=201)
async def create_expense(body: ExpenseCreate, db: DBSession, current_user: CurrentUser):
    if current_user["role"] not in ("admin", "operator"):
        raise HTTPException(status_code=403, detail="无权录入支出")

    data = body.model_dump()
    data.pop("floor", None)  # floor 仅入参用于解析范围，不是 Expense 列
    await _resolve_expense_scope(db, data, floor=body.floor)

    # 自住单关联的任何费用都由公司承担，不能靠前端 payer 输入绕过。
    if data.get("order_id"):
        linked_order = await db.scalar(
            select(Order).where(Order.order_id == data["order_id"])
        )
        if linked_order is not None and is_owner_self_order(linked_order):
            data["payer"] = ExpensePayer.company

    expense = Expense(
        expense_id="EXP-" + uuid.uuid4().hex[:12].upper(),
        created_by=current_user["user_id"],
        **data,
    )
    db.add(expense)
    await log_action_tx(db, current_user["user_id"], "expense.create", "expense", expense.expense_id)
    await db.commit()
    await db.refresh(expense)
    return expense


@router.post("/cleaning-charges/renewal", response_model=ExpenseOut, status_code=201)
async def create_renewal_cleaning_charge(
    body: RenewalCleaningChargeCreate,
    db: DBSession,
    current_user: CurrentUser,
    background_tasks: BackgroundTasks,
):
    """续住/在住房打扫登记 → 建一条保洁费用（收房东 30）。

    王总 2026-07-12 拍板：前台登记即计费（信任前台），金额服务端固定，房东全担。
    红线：只记账，绝不改房间状态（房里还住着人）。退房打扫走另一条自动链路。
    """
    if current_user["role"] not in ("admin", "operator", "keeper"):
        raise HTTPException(status_code=403, detail="无权登记打扫")

    order = (
        await db.execute(
            select(Order).where(Order.order_id == body.order_id, Order.is_deleted == False)
        )
    ).scalar_one_or_none()
    if not order:
        raise HTTPException(status_code=404, detail="订单不存在")

    room = (
        await db.execute(select(Room).where(Room.room_id == body.room_id))
    ).scalar_one_or_none()
    if not room:
        raise HTTPException(status_code=404, detail="房间不存在")
    if not room.owner_id:
        raise HTTPException(status_code=400, detail="该房间未配置业主，无法登记保洁费")

    # 房间必须属于该订单，否则费用挂到无关订单坏审计溯源（评审 Finding 4）。
    # 兼容老单只有 orders.room_id 单字段、未拆 OrderRoom 的情况。
    in_order = (
        await db.execute(
            select(OrderRoom.order_room_id)
            .where(OrderRoom.order_id == order.order_id, OrderRoom.room_id == room.room_id)
            .limit(1)
        )
    ).scalar_one_or_none()
    if in_order is None and order.room_id != room.room_id:
        raise HTTPException(status_code=400, detail="该房间不属于此订单")

    # 防重复：同订单同房同一天已登记过续住打扫则拒绝（前台双击 / 客户端重试）。评审 Finding 2。
    dup = (
        await db.execute(
            select(Expense.expense_id)
            .where(
                Expense.order_id == order.order_id,
                Expense.room_id == room.room_id,
                Expense.category == ExpenseCategory.cleaning,
                Expense.expense_date == today_cn(),
                Expense.is_deleted == False,
                Expense.description.like("续住打扫%"),
            )
            .limit(1)
        )
    ).scalar_one_or_none()
    if dup is not None:
        raise HTTPException(
            status_code=409,
            detail="今日该房已登记续住打扫，如确需再记请先作废原记录",
        )

    fees = await get_service_fees(db)
    from app.core.cleaning_pricing import OWNER_SELF_INSTAY_CLEANING_FEE
    owner_self = is_owner_self_order(order)
    expense = Expense(
        expense_id="EXP-" + uuid.uuid4().hex[:12].upper(),
        category=ExpenseCategory.cleaning,
        amount=OWNER_SELF_INSTAY_CLEANING_FEE if owner_self else fees.instay_cleaning_fee,
        description=f"续住打扫 {room.room_name}",
        expense_date=today_cn(),
        room_id=room.room_id,
        order_id=order.order_id,
        payer=ExpensePayer.company if owner_self else ExpensePayer.owner,
        owner_id=room.owner_id,
        is_service_fee=True,
        notes=body.notes,
        created_by=current_user["user_id"],
    )
    db.add(expense)
    await log_action_tx(
        db, current_user["user_id"], "expense.cleaning_renewal", "expense", expense.expense_id,
        notes=f"续住打扫登记 order={order.order_id} room={room.room_id}",
    )
    # 提交前把通知要用的值取成本地字符串（commit 后 ORM 属性可能过期）。
    _room_name = room.room_name
    _guest_name = order.guest_name
    _stay_range = f"{order.check_in_date.strftime('%m-%d')} → {order.check_out_date.strftime('%m-%d')}"
    await db.commit()
    await db.refresh(expense)

    # 王总确认要：登记时给保洁群发一张告知卡（best-effort，飞书挂了不影响登记）。
    # 放后台任务里，前台响应不被飞书往返拖慢。
    background_tasks.add_task(
        send_cleaning_renewal_alert,
        room_name=_room_name,
        guest_name=_guest_name,
        stay_range=_stay_range,
        notes=body.notes,
    )
    return expense


@router.delete("/cleaning-charges/{expense_id}", status_code=204)
async def void_cleaning_charge(expense_id: str, db: DBSession, current_user: CurrentUser):
    """作废一条保洁费用（软删）——前台登记错了自己撤。

    只作废「保洁」类目，别的支出走 /expenses 的管理员删除。软删后从结算聚合剔除
    （结算读 is_deleted==False）；已确认/打款的历史结算已快照，不受影响。
    """
    if current_user["role"] not in ("admin", "operator", "keeper"):
        raise HTTPException(status_code=403, detail="无权作废保洁费")

    expense = (
        await db.execute(
            select(Expense).where(
                Expense.expense_id == expense_id, Expense.is_deleted == False
            )
        )
    ).scalar_one_or_none()
    if not expense:
        raise HTTPException(status_code=404, detail="保洁费记录不存在")
    if expense.category != ExpenseCategory.cleaning:
        raise HTTPException(status_code=400, detail="该记录不是保洁费，不能在此作废")

    expense = await _reread_automatic_fee_after_owner_lock(db, expense)
    if not expense:
        raise HTTPException(status_code=404, detail="保洁费记录不存在")
    if expense.category != ExpenseCategory.cleaning:
        raise HTTPException(status_code=400, detail="该记录不是保洁费，不能在此作废")

    before = {"amount": str(expense.amount), "room_id": expense.room_id,
              "order_id": expense.order_id}
    expense.is_deleted = True
    expense.deleted_at = datetime.now(timezone.utc)
    expense.deleted_by = current_user["user_id"]
    await db.flush()
    await log_action_tx(
        db, current_user["user_id"], "expense.cleaning_void", "expense", expense_id,
        before_data=before,
    )
    await db.commit()


@router.get("/expenses", response_model=list[ExpenseOut])
async def list_expenses(
    db: DBSession,
    current_user: CurrentUser,
    room_id: Optional[str] = Query(default=None),
    year: Optional[int] = Query(default=None),
    month: Optional[int] = Query(default=None),
    start_date: Optional[date] = Query(default=None, description="起始日期 YYYY-MM-DD"),
    end_date: Optional[date] = Query(default=None, description="结束日期 YYYY-MM-DD"),
):
    """支出明细。日期过滤二选一（start_date+end_date 优先，向后兼容 year+month）。"""
    if current_user["role"] not in ("admin", "operator", "finance"):
        raise HTTPException(status_code=403, detail="无权查看支出")
    q = select(Expense).where(Expense.is_deleted == False)
    if room_id:
        q = q.where(Expense.room_id == room_id)
    if start_date and end_date:
        if end_date < start_date:
            raise HTTPException(status_code=400, detail="end_date 必须 >= start_date")
        q = q.where(Expense.expense_date >= start_date)
        q = q.where(Expense.expense_date <= end_date)
    elif year and month:
        from sqlalchemy import extract
        q = q.where(extract("year", Expense.expense_date) == year)
        q = q.where(extract("month", Expense.expense_date) == month)
    result = await db.execute(q.order_by(Expense.expense_date.desc()))
    return result.scalars().all()


@router.patch("/expenses/{expense_id}", response_model=ExpenseOut)
async def update_expense(expense_id: str, body: ExpenseCreate, db: DBSession, current_user: CurrentUser):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="仅管理员可修改支出")
    result = await db.execute(
        select(Expense).where(Expense.expense_id == expense_id, Expense.is_deleted == False)
    )
    expense = result.scalar_one_or_none()
    if not expense:
        raise HTTPException(status_code=404, detail="支出记录不存在")
    if expense.is_service_fee:
        raise HTTPException(
            status_code=409,
            detail="自动服务费不可直接修改；请先作废/删除该记录，再由服务费对账重新生成",
        )
    before = {"amount": str(expense.amount), "category": expense.category, "description": expense.description}
    patch = body.model_dump(exclude_none=True)
    patch.pop("floor", None)  # floor 非 Expense 列，编辑时忽略
    for field, value in patch.items():
        setattr(expense, field, value)
    await db.flush()
    after = {"amount": str(expense.amount), "category": expense.category, "description": expense.description}
    await log_action_tx(
        db, current_user["user_id"], "expense.update", "expense", expense_id,
        before_data=before, after_data=after,
    )
    await db.commit()
    await db.refresh(expense)
    return expense


@router.delete("/expenses/{expense_id}", status_code=204)
async def delete_expense(expense_id: str, db: DBSession, current_user: CurrentUser):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="仅管理员可删除支出")
    result = await db.execute(
        select(Expense).where(Expense.expense_id == expense_id, Expense.is_deleted == False)
    )
    expense = result.scalar_one_or_none()
    if not expense:
        raise HTTPException(status_code=404, detail="支出记录不存在")

    expense = await _reread_automatic_fee_after_owner_lock(db, expense)
    if not expense:
        raise HTTPException(status_code=404, detail="支出记录不存在")

    # 软删除（批2 item7）：保留行做审计 + 不破坏历史分账溯源；从所有聚合中剔除。
    before = {"amount": str(expense.amount), "category": expense.category,
              "description": expense.description, "room_id": expense.room_id}
    expense.is_deleted = True
    expense.deleted_at = datetime.now(timezone.utc)
    expense.deleted_by = current_user["user_id"]
    await db.flush()
    await log_action_tx(
        db, current_user["user_id"], "expense.delete", "expense", expense_id,
        before_data=before,
    )
    await db.commit()


# ─── 业主服务费费率配置 ───────────────────────────────────────────────────

@router.get("/service-fee-config", response_model=ServiceFeeConfigOut)
async def get_service_fee_config(db: DBSession, current_user: CurrentUser):
    """读当前服务费费率（保洁/续住/洗涤/日耗）。缺行则按默认建一行。"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="仅管理员可查看费用标准")
    row = await get_or_create_config_row(db)
    await db.commit()
    return row


@router.put("/service-fee-config", response_model=ServiceFeeConfigOut)
async def update_service_fee_config(
    body: ServiceFeeConfigUpdate, db: DBSession, current_user: CurrentUser
):
    """改服务费费率——即时生效（下一单退房记账即按新价）。历史已记账支出不变。"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="仅管理员可修改费用标准")
    row = await get_or_create_config_row(db)
    before = {
        "checkout_cleaning_fee": str(row.checkout_cleaning_fee),
        "instay_cleaning_fee": str(row.instay_cleaning_fee),
        "laundry_fee_per_room": str(row.laundry_fee_per_room),
        "consumable_fee_per_room_night": str(row.consumable_fee_per_room_night),
    }
    row.checkout_cleaning_fee = body.checkout_cleaning_fee
    row.instay_cleaning_fee = body.instay_cleaning_fee
    row.laundry_fee_per_room = body.laundry_fee_per_room
    row.consumable_fee_per_room_night = body.consumable_fee_per_room_night
    row.updated_by = current_user["user_id"]
    await db.flush()
    after = {
        "checkout_cleaning_fee": str(row.checkout_cleaning_fee),
        "instay_cleaning_fee": str(row.instay_cleaning_fee),
        "laundry_fee_per_room": str(row.laundry_fee_per_room),
        "consumable_fee_per_room_night": str(row.consumable_fee_per_room_night),
    }
    await log_action_tx(
        db, current_user["user_id"], "service_fee_config.update",
        "service_fee_config", "default", before_data=before, after_data=after,
    )
    await db.commit()
    await db.refresh(row)
    return row


# ─── Summary by Room ─────────────────────────────────────────────────────────

@router.get("/summary/by-room", response_model=list[ByRoomSummaryItem])
async def summary_by_room(
    db: DBSession,
    current_user: CurrentUser,
    year: Optional[int] = Query(default=None),
    month: Optional[int] = Query(default=None, ge=1, le=12),
    start_date: Optional[date] = Query(default=None, description="起始日期 YYYY-MM-DD"),
    end_date: Optional[date] = Query(default=None, description="结束日期 YYYY-MM-DD"),
):
    """
    按房号聚合财务。日期范围可通过两种方式指定(二选一):
      1) start_date + end_date : 任意日期范围 (优先使用)
      2) year + month           : 单月 (向后兼容)

    聚合字段:
      revenue        = SUM(orders.actual_price) WHERE check_out_date IN range
      commission     = SUM(revenue * platform_commission_rate)
      net_revenue    = revenue - commission + OTA平台补贴
      expense_company= SUM(expenses.amount) WHERE payer=company
      expense_owner  = SUM(expenses.amount) WHERE payer=owner
      owner_net_share= net_revenue * owner_share_ratio - 按规则加权后的业主支出
    """
    if current_user["role"] not in ("admin", "operator", "finance", "owner"):
        raise HTTPException(status_code=403, detail="无权查看分账")

    # 把 year/month 也统一成 start_date/end_date
    if not (start_date and end_date):
        if not (year and month):
            raise HTTPException(status_code=400, detail="请提供 start_date+end_date 或 year+month")
        from calendar import monthrange
        _, days = monthrange(year, month)
        start_date = date(year, month, 1)
        end_date = date(year, month, days)
    if end_date < start_date:
        raise HTTPException(status_code=400, detail="end_date 必须 >= start_date")

    # Expenses per room, split by payer
    exp_q = (
        select(
            Expense.room_id.label("room_id"),
            func.coalesce(
                func.sum(case((Expense.payer == ExpensePayer.company, Expense.amount), else_=0)),
                0,
            ).label("expense_company"),
            func.coalesce(
                func.sum(case((Expense.payer == ExpensePayer.owner, Expense.amount), else_=0)),
                0,
            ).label("expense_owner"),
        )
        .where(Expense.room_id.isnot(None))
        .where(Expense.is_deleted == False)
        .where(Expense.expense_date >= start_date)
        .where(Expense.expense_date <= end_date)
        .group_by(Expense.room_id)
    )
    exp_rows = {r.room_id: r for r in (await db.execute(exp_q)).all()}

    # All rooms for stable ordering + owner info
    rooms_result = await db.execute(select(Room).order_by(Room.room_id))
    rooms = rooms_result.scalars().all()

    # 唯一正式口径：补贴、单房 OTA 原始到手、RoomCostShareRule 和 v2 分成公式
    # 全部复用结算服务。按房号页不再维护第二套近似公式。
    from app.services.owner_settlement import compute_room_period_owner_stat

    out: list[ByRoomSummaryItem] = []
    for room in rooms:
        stat = await compute_room_period_owner_stat(db, room, start_date, end_date)
        exp = exp_rows.get(room.room_id)
        revenue = stat.revenue
        commission = stat.commission
        order_count = stat.order_count
        expense_company = Decimal(exp.expense_company) if exp else Decimal("0")
        expense_owner = stat.owner_expenses
        net_revenue = stat.net_revenue
        share_ratio = stat.share_ratio
        owner_net_share = stat.owner_net

        out.append(
            ByRoomSummaryItem(
                room_id=room.room_id,
                room_name=room.room_name,
                owner_id=room.owner_id,
                owner_share_ratio=share_ratio,
                order_count=order_count,
                revenue=revenue.quantize(Decimal("0.01")),
                commission=commission.quantize(Decimal("0.01")),
                net_revenue=net_revenue.quantize(Decimal("0.01")),
                expense_company=expense_company.quantize(Decimal("0.01")),
                expense_owner=expense_owner.quantize(Decimal("0.01")),
                owner_net_share=owner_net_share,
                company_net=(net_revenue - expense_company - owner_net_share).quantize(Decimal("0.01")),
            )
        )
    return out


@router.get("/summary/room-segments", response_model=list[RoomRevenueSegment])
async def summary_room_segments(
    db: DBSession,
    current_user: CurrentUser,
    room_id: str = Query(..., description="房号"),
    year: Optional[int] = Query(default=None),
    month: Optional[int] = Query(default=None, ge=1, le=12),
    start_date: Optional[date] = Query(default=None, description="起始日期 YYYY-MM-DD"),
    end_date: Optional[date] = Query(default=None, description="结束日期 YYYY-MM-DD"),
):
    """单房详情页「区间订单（收入来源）」明细：按 OrderRoom **段**返回该房的营收行。

    为什么不是 GET /orders?room_id=：那条路走 Order.room_id 顶层字段过滤，而多房订单
    顶层 room_id 为空（房号落在各 OrderRoom 上）→ 多房单不会出现在单房明细里，明细少行，
    与上方「区间收入」汇总卡（summary_by_room 按 OrderRoom 聚合）对不上，看着像漏单。

    口径与 summary_by_room 的营收行**逐条一致**（同一份 WHERE）：
      - Order.is_deleted == False 且 order_status != cancelled
      - OrderRoom.room_id == room_id
      - OrderRoom.check_out_date ∈ [start, end]（收入按离店日归属，2026-07-14 拍板）
      - 实收取段级 OrderRoom.actual_price（多房单每段独立），不是订单顶层 actual_price
    段级实收合计 == 该房 summary_by_room.revenue。
    """
    if current_user["role"] not in ("admin", "operator", "finance", "owner"):
        raise HTTPException(status_code=403, detail="无权查看分账")

    # 把 year/month 也统一成 start_date/end_date（与 summary_by_room 同款兜底）
    if not (start_date and end_date):
        if not (year and month):
            raise HTTPException(status_code=400, detail="请提供 start_date+end_date 或 year+month")
        from calendar import monthrange
        _, days = monthrange(year, month)
        start_date = date(year, month, 1)
        end_date = date(year, month, days)
    if end_date < start_date:
        raise HTTPException(status_code=400, detail="end_date 必须 >= start_date")

    from app.models.order_room import OrderRoom
    rows = (await db.execute(
        select(
            OrderRoom.order_id, OrderRoom.order_room_id, OrderRoom.room_id,
            OrderRoom.check_in_date, OrderRoom.check_out_date, OrderRoom.actual_price,
            Order.guest_name, Order.channel,
        )
        .join(Order, Order.order_id == OrderRoom.order_id)
        .where(Order.is_deleted == False)
        .where(Order.order_status != OrderStatus.cancelled)
        .where(OrderRoom.room_id == room_id)
        .where(OrderRoom.check_out_date >= start_date)
        .where(OrderRoom.check_out_date <= end_date)
        # 入住日升序，同日按段键稳定（翻页/多次查询顺序一致）
        .order_by(OrderRoom.check_in_date.asc(), OrderRoom.order_room_id.asc())
    )).all()

    return [
        RoomRevenueSegment(
            order_id=r.order_id,
            order_room_id=r.order_room_id,
            room_id=r.room_id,
            guest_name=r.guest_name,
            channel=r.channel,
            check_in_date=r.check_in_date,
            check_out_date=r.check_out_date,
            nights=(r.check_out_date - r.check_in_date).days,
            actual_price=r.actual_price,
        )
        for r in rows
    ]


# ─── Expense Excel Import ────────────────────────────────────────────────────

# 模板脚手架行（示例数据 + 底部「说明:」帮助文案块）的隐藏标记，写在第 8 列 (H)。
# 导入时据此跳过这些行 —— 避免用户下载模板原样上传时，把 3 行示例当真实支出入库、
# 把只填第 1 列的说明块当失败行报告。第 8 列在模板里被隐藏，用户可见的仍是 7 列。
_TEMPLATE_MARKER = "__TEMPLATE_SAMPLE__"
_TEMPLATE_MARKER_COL = 8                       # openpyxl 1-based 列号 (H)
_TEMPLATE_MARKER_IDX = _TEMPLATE_MARKER_COL - 1  # values_only 元组 0-based 下标

# 中文精确 label → enum value（含"（旧）"去后缀变体）。由枚举权威 label 反推，
# 不手抄，避免漂移。精确匹配在模糊子串之前跑，规避"公摊水费"含"水费"这类碰撞。
_LABEL_TO_VALUE: dict[str, str] = {}
for _cat, _label in EXPENSE_CATEGORY_LABELS.items():
    _LABEL_TO_VALUE[_label] = _cat.value
    _base = _label.replace("（旧）", "").replace("(旧)", "")
    _LABEL_TO_VALUE.setdefault(_base, _cat.value)

# 模糊子串兜底（用户随手写"水电""佣金"等）。顺序 = 具体优先，避免子串误命中：
# 厨房保洁/续住保洁 先于 保洁；公摊水费/新布草过水 先于 水费；物业引导 先于 物业费。
_CATEGORY_ZH = {
    "厨房保洁": "kitchen_cleaning",
    "续住保洁": "cleaning",
    "保洁": "cleaning",
    "维修": "maintenance",
    "公摊水费": "public_utilities",
    "公摊": "public_utilities",
    "新布草过水": "new_linen_prewash",
    "新布草": "new_linen_prewash",
    "水费": "water",
    "水电煤": "utilities",
    "水电": "utilities",
    "冷水": "cold_water",
    "热水": "hot_water",
    "电费": "electricity",
    "燃气": "gas",
    "宽带": "broadband",
    "日耗": "daily_supplies",
    "洗涤": "laundry",
    "物业引导": "property_guidance_fee",
    "物业": "property_fee",
    "采购": "supplies",
    "物品采购": "supplies",
    "物品": "supplies",
    "平台手续费": "platform_fee",
    "平台": "platform_fee",
    "佣金": "platform_fee",
    "税费": "tax",
    "税": "tax",
    "其他": "other",
}

# 导入白名单从枚举派生（含停用值，历史英文数据仍可导入），不随新增类目漂移。
_VALID_CATEGORY_VALUES = {c.value for c in ExpenseCategory}


def _normalize_category(raw: str) -> Optional[str]:
    if not raw:
        return None
    s = str(raw).strip()
    if s.lower() in _VALID_CATEGORY_VALUES:  # 英文 enum 原值
        return s.lower()
    if s in _LABEL_TO_VALUE:                 # 中文精确 label
        return _LABEL_TO_VALUE[s]
    for zh, en in _CATEGORY_ZH.items():      # 中文模糊子串兜底
        if zh in s:
            return en
    return None


def _normalize_payer(raw: str) -> str:
    if not raw:
        return "company"
    s = str(raw).strip()
    if s in ("owner", "业主", "房东"):
        return "owner"
    return "company"


def _normalize_room_id(raw, valid_ids: set[str]) -> Optional[str]:
    """模糊匹配房号: '101' 和 'R101' 都能认"""
    if raw is None or str(raw).strip() == "":
        return None
    s = str(raw).strip().upper()
    if s in valid_ids:
        return s
    # 去掉 R/房间/Room 等前缀再试
    stripped = re.sub(r"^(R|房间|ROOM\s*)", "", s).strip()
    if stripped in valid_ids:
        return stripped
    # 加 R 前缀再试
    if f"R{stripped}" in valid_ids:
        return f"R{stripped}"
    return None


def _parse_date(raw) -> Optional[date]:
    if raw is None or str(raw).strip() == "":
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    s = str(raw).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


@router.get("/expenses/import-template")
async def expense_import_template(current_user: CurrentUser):
    """下载支出导入模板 Excel"""
    if current_user["role"] not in ("admin", "operator", "finance"):
        raise HTTPException(status_code=403, detail="无权下载模板")

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "支出导入模板"

    headers = ["日期", "类别", "金额", "描述", "关联房号", "支付方", "备注"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2E5CFF")
        cell.alignment = Alignment(horizontal="center")

    # 示例行
    examples = [
        ["2026-04-10", "保洁", 80, "王阿姨 101 房清洁", "101", "公司", ""],
        ["2026-04-12", "维修", 320, "空调维修", "R102", "业主", "发票 #A001"],
        ["2026-04-15", "水电煤", 180, "全屋电费", "", "公司", ""],
    ]
    for r, row in enumerate(examples, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
        # 隐藏标记：导入时据此跳过示例行，防止原样上传把示例当真实支出
        ws.cell(row=r, column=_TEMPLATE_MARKER_COL, value=_TEMPLATE_MARKER)

    # 说明行
    note_row = len(examples) + 3
    ws.cell(row=note_row, column=1, value="说明:").font = Font(bold=True)
    notes = [
        "1. 日期格式: YYYY-MM-DD 或 YYYY/MM/DD",
        "2. 类别可填中文或英文 (保洁/cleaning, 维修/maintenance, 水电煤/utilities, 物品采购/supplies, 平台手续费/platform_fee, 税费/tax, 其他/other)",
        "3. 关联房号可选；支持「101」或「R101」",
        "4. 支付方: 公司 或 业主 (业主承担的将从业主分成中扣除)",
        "5. 导入前请删除以上 3 行示例；未删除也会被自动忽略，不会入库。",
    ]
    # 说明块每行都打隐藏标记：导入时整块跳过，不计入失败明细
    ws.cell(row=note_row, column=_TEMPLATE_MARKER_COL, value=_TEMPLATE_MARKER)
    for i, n in enumerate(notes):
        ws.cell(row=note_row + 1 + i, column=1, value=n)
        ws.cell(row=note_row + 1 + i, column=_TEMPLATE_MARKER_COL, value=_TEMPLATE_MARKER)

    # 列宽
    widths = [12, 12, 10, 30, 12, 10, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    # 隐藏标记列，用户不可见
    ws.column_dimensions[openpyxl.utils.get_column_letter(_TEMPLATE_MARKER_COL)].hidden = True

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="expense-template.xlsx"'},
    )


@router.post("/expenses/import", response_model=ExpenseImportResult)
async def import_expenses(
    db: DBSession,
    current_user: CurrentUser,
    file: UploadFile = File(...),
):
    """
    从 Excel 批量导入支出。成功行入库，失败行返回错误不阻塞。
    """
    if current_user["role"] not in ("admin", "operator", "finance"):
        raise HTTPException(status_code=403, detail="无权导入支出")

    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 文件")

    import openpyxl

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"无法读取 Excel: {e}")
    ws = wb.active

    # 预加载所有有效房号用于模糊匹配
    rooms_result = await db.execute(select(Room.room_id))
    valid_room_ids = {r for (r,) in rooms_result.all()}

    succeeded: list[Expense] = []
    failed: list[dict] = []
    total_rows = 0

    # 跳过表头第 1 行
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        cells = list(row)
        # 跳过模板脚手架行（示例行 + 说明块）：靠隐藏标记列识别。
        # 原样上传的模板不会产生任何入库数据，也不进失败明细。
        marker = cells[_TEMPLATE_MARKER_IDX] if len(cells) > _TEMPLATE_MARKER_IDX else None
        if marker is not None and str(marker).strip() == _TEMPLATE_MARKER:
            continue
        # 跳过完全空行
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        # 跳过「仅第 1 列有值」的行（说明块残留或误填）：缺类别/金额/描述，
        # 不可能是有效支出，不计入失败明细，避免污染批量结果。
        if all(v is None or str(v).strip() == "" for v in cells[1:7]):
            continue
        total_rows += 1
        try:
            raw_date, raw_category, raw_amount, raw_desc, raw_room, raw_payer, raw_notes = (
                list(row) + [None] * 7
            )[:7]

            errors: list[str] = []
            expense_date = _parse_date(raw_date)
            if not expense_date:
                errors.append("日期格式错误")

            category = _normalize_category(raw_category)
            if not category:
                errors.append(f"未知类别: {raw_category}")

            try:
                amount = Decimal(str(raw_amount).strip())
                if amount <= 0:
                    errors.append("金额必须大于 0")
            except (InvalidOperation, AttributeError, TypeError):
                amount = Decimal("0")
                errors.append(f"金额格式错误: {raw_amount}")

            description = str(raw_desc).strip() if raw_desc else ""
            if not description:
                errors.append("描述不能为空")

            room_id = _normalize_room_id(raw_room, valid_room_ids) if raw_room else None
            if raw_room and not room_id:
                errors.append(f"房号未找到: {raw_room}")

            payer = _normalize_payer(raw_payer)
            notes = str(raw_notes).strip() if raw_notes else None

            if errors:
                failed.append({
                    "row": row_idx,
                    "errors": "; ".join(errors),
                    "data": {
                        "date": str(raw_date), "category": str(raw_category),
                        "amount": str(raw_amount), "description": str(raw_desc),
                        "room": str(raw_room), "payer": str(raw_payer),
                    },
                })
                continue

            expense = Expense(
                expense_id="EXP-" + uuid.uuid4().hex[:12].upper(),
                category=category,
                amount=amount,
                description=description[:200],
                expense_date=expense_date,
                room_id=room_id,
                payer=payer,
                notes=notes[:200] if notes else None,
                created_by=current_user["user_id"],
            )
            db.add(expense)
            succeeded.append(expense)
        except Exception as e:
            failed.append({
                "row": row_idx,
                "errors": f"解析失败: {e}",
                "data": {},
            })

    if succeeded:
        await log_action_tx(
            db, current_user["user_id"], "expense.import",
            "expense", f"imported {len(succeeded)} rows",
        )
        await db.commit()
        for exp in succeeded:
            await db.refresh(exp)

    return ExpenseImportResult(
        succeeded=succeeded,
        failed=failed,
        total_rows=total_rows,
        imported_count=len(succeeded),
    )
