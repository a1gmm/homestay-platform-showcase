# backend/app/services/billing_recon/parser.py
"""账单表格 → 结构化行。AI 只产出 BillMapping（认列），本文件全部确定性代码。

校验闸（validate_bill）只有两类硬闸，errors 非空 = 整批拒收：
①逐行加总 != 账单汇总金额（容差 0.01）；
②窗口内行占比 < 80%（窗口 = 账单月 ±7 天；窗口外行本身合法——跨月补结/退款——
  不逐行拒收，只在占比过低时判定 AI 大概率认错列/认错表）。
stats 里如实记录 out_of_window / in_window_ratio，供落库展示，不参与拒收判断之外的用途。
"""
from __future__ import annotations

import io
import logging
import zipfile
from collections import Counter
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Annotated, Literal

from pydantic import BaseModel, Field

logger = logging.getLogger(__name__)

_D01 = Decimal("0.01")
_DATE_FORMATS = ("%Y/%m/%d", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M:%S")
_WINDOW_RATIO_GATE = 0.8  # 窗口内占比 < 此值 -> 整批拒收
# AI 输出信任边界：单表最多物化多少行/列。行超限 = 整批拒收（多半是损坏文件/炸弹表），
# 列只截断（真实账单不会有 64 列以上的有效数据，多出来的都是格式垃圾）。
_MAX_SHEET_ROWS = 20000
_MAX_SHEET_COLS = 64


class BillParseError(ValueError):
    """账单文件本身打不开/不是合法 xls(x)：整批拒收，不进入解析。"""


# 列号/行号一律非负：AI 吐负数会被 Python 的负索引悄悄绕回行尾读到别的列（静默错数），
# 必须在信任边界上直接拒收，而不是靠下游 _cell 兜底。
NonNegInt = Annotated[int, Field(ge=0)]
RowTypeName = Literal["normal", "refund", "compensation"]


class BillMapping(BaseModel):
    sheet: str
    header_row: NonNegInt
    col_order_no: NonNegInt
    col_guest: NonNegInt
    col_checkin: NonNegInt
    col_checkout: NonNegInt
    col_amount: NonNegInt
    col_row_type: NonNegInt | None = None
    # 只认引擎认识的三种行类型；AI 编出别的值（如 "cancel"）直接 ValidationError，
    # 不能让它悄悄落成一个下游谁也不处理的 row_type。
    row_type_map: dict[str, RowTypeName] = {}
    summary_total: float = Field(allow_inf_nan=False)
    platform_guess: str = "ctrip"


@dataclass
class BillRow:
    order_no: str
    guest: str
    checkin: date | None
    checkout: date | None
    amount: Decimal
    row_type: str  # normal | refund | compensation
    # 金额单元格原文非空、但解析不出数字（被当 0 处理）——可见性标记，
    # validate_bill 汇总进 stats["unparsed_amounts"]，总额闸失败时附在错误信息里。
    amount_unparsed: bool = False


@dataclass
class BillOrder:
    order_no: str
    guest: str
    checkout: date | None
    net: Decimal
    row_types: set[str] = field(default_factory=set)
    # 赔款(compensation)类型行金额合计（负数=账单扣款）。0 = 该单没有赔款行。
    # 供 Task 5 引擎 classify() 在“匹配上系统单但账单里混了一笔赔款调整”场景下
    # 落 detail.has_compensation/compensation_amount（终稿 R3），供前端/Task 9 展示。
    compensation_amount: Decimal = Decimal("0")


def load_workbook_rows(data: bytes, filename: str) -> tuple[dict[str, list[list]], int]:
    """xls/xlsx → ({sheet名: 行列表}, datemode)。

    xls 走 xlrd（携程账单是老格式），datemode 取 wb.datemode（修 Mac 1904 纪元偏 4 年的坑）；
    xlsx 走 openpyxl，datemode 恒为 0（openpyxl 直接给 datetime，不需要按纪元换算）。
    打不开的坏文件统一包装成 BillParseError（底层库异常五花八门——xlrd 的 CompDocError、
    openpyxl 的 KeyError/ValueError 等——一律兜住，调用方不用分别识别）。
    行/列上限见 _MAX_SHEET_ROWS/_MAX_SHEET_COLS：行超限拒收，列只截断。
    """
    name = (filename or "").lower()
    if name.endswith(".xls"):
        import xlrd

        try:
            wb = xlrd.open_workbook(file_contents=data)
        except Exception as e:  # noqa: BLE001 — 坏 xls 会抛 XLRDError/CompDocError/struct.error 等
            raise BillParseError(f"无法解析 xls 文件: {e}") from e
        try:
            sheets = {}
            for sh in wb.sheets():
                if sh.nrows > _MAX_SHEET_ROWS:
                    raise BillParseError("表格行数超限(>20000)，请确认不是损坏文件")
                ncols = min(sh.ncols, _MAX_SHEET_COLS)
                sheets[sh.name] = [[sh.cell_value(r, c) for c in range(ncols)] for r in range(sh.nrows)]
        except BillParseError:
            raise
        except Exception as e:  # noqa: BLE001 — 截断的 xls 读到一半才炸
            raise BillParseError(f"无法解析 xls 文件: {e}") from e
        return sheets, wb.datemode

    import openpyxl
    from openpyxl.utils.exceptions import InvalidFileException

    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except (InvalidFileException, zipfile.BadZipFile) as e:
        raise BillParseError(f"无法解析 xlsx 文件: {e}") from e
    except Exception as e:  # noqa: BLE001 — 畸形 zip 内容会抛 KeyError/ValueError 等
        raise BillParseError(f"无法解析 xlsx 文件: {e}") from e
    try:
        sheets = {}
        for ws in wb.worksheets:
            rows: list[list] = []
            for row in ws.iter_rows(values_only=True):
                if len(rows) >= _MAX_SHEET_ROWS:
                    raise BillParseError("表格行数超限(>20000)，请确认不是损坏文件")
                rows.append(list(row[:_MAX_SHEET_COLS]))
            sheets[ws.title] = rows
    except BillParseError:
        raise
    except Exception as e:  # noqa: BLE001 — read_only 流式读到坏 sheet 才炸
        raise BillParseError(f"无法解析 xlsx 文件: {e}") from e
    return sheets, 0


def _cell(row: list, idx: int | None):
    # idx < 0 一律当空：Python 负索引会绕回行尾读到完全无关的列，静默错数比读空危险得多。
    if idx is None or idx < 0 or idx >= len(row) or row[idx] is None:
        return ""
    return row[idx]


def _dec_checked(v) -> tuple[Decimal, bool]:
    """→ (金额, 是否"原文非空但解析不出")。空单元格不算 unparsed（正常留白）。"""
    s = str(v).replace(",", "").replace("¥", "").strip()
    if not s:
        return Decimal("0"), False
    try:
        return Decimal(s).quantize(_D01), False
    except InvalidOperation:
        return Decimal("0"), True


def _dec(v) -> Decimal:
    return _dec_checked(v)[0]


def _parse_date(v, datemode: int = 0) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)) and v > 0:
        import xlrd

        try:
            return xlrd.xldate_as_datetime(v, datemode).date()
        except Exception:
            return None
    s = str(v).strip()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _order_no(v) -> str | None:
    """单号列取值 → 纯数字串；xls 里长数字可能是 float。非订单行（表头/合计）返回 None。"""
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    s = s.split(".")[0]
    return s if s.isdigit() and len(s) >= 8 else None


def extract_bill_rows(rows: list[list], m: BillMapping, datemode: int = 0) -> list[BillRow]:
    if m.header_row >= len(rows):
        raise BillParseError("表头行号越界")
    out: list[BillRow] = []
    for i, r in enumerate(rows[m.header_row + 1 :], start=m.header_row + 1):
        raw_no = _cell(r, m.col_order_no)
        no = _order_no(raw_no)
        if no is None:
            # 丢行可见性：账单里的标题/表头/合计行本来就该丢，但 AI 认错 col_order_no 时
            # 也长这样——服务端留痕行号+原文，排查时一眼看出是"整表都在丢"还是只丢了几行。
            if str(raw_no).strip():
                logger.warning("billing_recon 跳过非订单行 row=%d cell=%r", i, raw_no)
            continue
        row_type = "normal"
        if m.col_row_type is not None:
            row_type = m.row_type_map.get(str(_cell(r, m.col_row_type)).strip(), "normal")
        amount, unparsed = _dec_checked(_cell(r, m.col_amount))
        out.append(
            BillRow(
                order_no=no,
                guest=str(_cell(r, m.col_guest)).strip(),
                checkin=_parse_date(_cell(r, m.col_checkin), datemode),
                checkout=_parse_date(_cell(r, m.col_checkout), datemode),
                amount=amount,
                row_type=row_type,
                amount_unparsed=unparsed,
            )
        )
    return out


def aggregate_orders(rows: list[BillRow]) -> dict[str, BillOrder]:
    out: dict[str, BillOrder] = {}
    for r in rows:
        o = out.get(r.order_no)
        if o is None:
            o = out[r.order_no] = BillOrder(order_no=r.order_no, guest=r.guest, checkout=r.checkout, net=Decimal("0"))
        o.net = (o.net + r.amount).quantize(_D01)
        o.row_types.add(r.row_type)
        if r.row_type == "compensation":
            o.compensation_amount = (o.compensation_amount + r.amount).quantize(_D01)
        if r.checkout and (o.checkout is None or r.checkout > o.checkout):
            o.checkout = r.checkout
    return out


def infer_bill_month(rows: list[BillRow]) -> str:
    months = Counter(r.checkout.strftime("%Y-%m") for r in rows if r.checkout)
    if not months:
        raise ValueError("账单里没有可解析的离店日期")
    return months.most_common(1)[0][0]


def _month_window(bill_month: str) -> tuple[date, date]:
    from calendar import monthrange
    from datetime import timedelta

    y, mo = int(bill_month[:4]), int(bill_month[5:7])
    first = date(y, mo, 1)
    last = date(y, mo, monthrange(y, mo)[1])
    return first - timedelta(days=7), last + timedelta(days=7)


def validate_bill(rows: list[BillRow], m: BillMapping) -> tuple[list[str], dict]:
    """两类硬闸；软闸只落 stats 不拦。errors 非空 = 整批拒收。"""
    errors: list[str] = []
    stats: dict = {"out_of_window": 0, "in_window_ratio": 1.0, "unparsed_amounts": 0}
    if not rows:
        return ["没有解析出任何明细行（AI 可能认错了表/列，请检查文件后重传）"], stats

    unparsed = sum(1 for r in rows if r.amount_unparsed)
    stats["unparsed_amounts"] = unparsed

    total = sum((r.amount for r in rows), Decimal("0")).quantize(_D01)
    try:
        expect = Decimal(str(m.summary_total)).quantize(_D01)
    except InvalidOperation:
        expect = None
    # nan/inf 进得了 Decimal 但比不出大小（NaN > 0.01 恒为 False），会让总额闸静默放行——
    # 必须显式判死。BillMapping 的 allow_inf_nan=False 是第一道，这里是绕过模型时的第二道。
    if expect is None or not expect.is_finite():
        # AI 吐出的汇总值不是个能比大小的数：整批拒收，不是 500，也不是静默通过
        errors.append(f"账单汇总金额不是合法数字：{m.summary_total!r}，整批拒收")
        expect = None
    if expect is not None and abs(total - expect) > _D01:
        msg = f"逐行加总 {total} ≠ 账单汇总 {expect}，整批拒收"
        if unparsed:
            # 总额对不上时最常见的根因就是有金额格没解析成功（AI 认错列/表里混了文字）
            msg += f"；另有 {unparsed} 个金额单元格无法解析"
        errors.append(msg)

    try:
        bill_month = infer_bill_month(rows)
    except ValueError as e:
        errors.append(str(e))
        return errors, stats

    lo, hi = _month_window(bill_month)
    # 分母 = 全部行（不只是能解析出 checkout 的行）：checkout 解析不出本身就是异常信号
    # （AI 可能把 col_checkout 认错到别的列），必须计入 out_of_window，不能被悄悄排除在分母外
    # 而让占比虚高、闸门裸奔。rows 非空已在函数开头保证，无需 len(rows)==0 兜底分支。
    out_of_window = [r for r in rows if r.checkout is None or not (lo <= r.checkout <= hi)]
    stats["out_of_window"] = len(out_of_window)
    stats["in_window_ratio"] = round((len(rows) - len(out_of_window)) / len(rows), 4)

    if stats["in_window_ratio"] < _WINDOW_RATIO_GATE:
        errors.append(
            f"{len(out_of_window)}/{len(rows)} 行离店日期超出账单月±7天窗口，"
            f"窗口内占比 {stats['in_window_ratio']:.0%} < 80%，整批拒收"
        )
    return errors, stats
