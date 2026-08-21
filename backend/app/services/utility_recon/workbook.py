"""有界读取异构 Excel，并确定文件角色、表头和月份范围。"""

from datetime import date, datetime
from io import BytesIO
import re
import zipfile

from openpyxl import load_workbook
import xlrd

from .contracts import DetectedTable, InspectedFile, PreflightResult, WorkbookInput


MAX_FILE_BYTES = 10 * 1024 * 1024
MAX_SHEETS = 50
MAX_ROWS = 20_000
MAX_COLUMNS = 100

_SYNONYMS = {
    "date": ("收款日期", "付款日期", "费用日期", "日期"),
    "floor": ("楼层", "楼栋楼层"),
    "room": ("房间号", "房号", "房间"),
    "customer": ("客户", "客户姓名", "住户", "姓名"),
    "category": ("费用科目", "科目", "费用类型", "项目"),
    "receipt_amount": ("已收金额", "收款金额", "实收金额"),
    "expense_amount": ("付款金额", "费用金额", "支出金额"),
    "summary": ("摘要", "备注", "说明"),
}


class WorkbookInspectionError(ValueError):
    pass


def _clean(value: object) -> str:
    return re.sub(r"\s+", "", str(value or "")).lower()


def _map_headers(row: tuple) -> dict[str, int]:
    mapped: dict[str, int] = {}
    for index, value in enumerate(row):
        cleaned = _clean(value)
        for field, synonyms in _SYNONYMS.items():
            if cleaned in {_clean(item) for item in synonyms}:
                mapped.setdefault(field, index)
    return mapped


def _role(columns: dict[str, int]) -> str | None:
    if "receipt_amount" in columns and "date" in columns:
        return "receipt"
    if "expense_amount" in columns and "date" in columns:
        return "expense"
    return None


def _month(value: object) -> str | None:
    if isinstance(value, (date, datetime)):
        return f"{value.year:04d}-{value.month:02d}"
    text = str(value or "").strip()
    match = re.search(r"(?P<year>20\d{2})\D{0,3}(?P<month>1[0-2]|0?[1-9])", text)
    if match:
        return f"{int(match.group('year')):04d}-{int(match.group('month')):02d}"
    return None


def _validate_input(item: WorkbookInput) -> None:
    suffix = item.filename.lower().rsplit(".", 1)[-1] if "." in item.filename else ""
    if suffix not in {"xls", "xlsx"}:
        raise WorkbookInspectionError("请上传 .xls 或 .xlsx Excel 文件")
    if len(item.content) > MAX_FILE_BYTES:
        raise WorkbookInspectionError("Excel 文件超过 10MB")
    if suffix == "xlsx" and not zipfile.is_zipfile(BytesIO(item.content)):
        raise WorkbookInspectionError("Excel 文件内容无效")


def _xlsx_rows(item: WorkbookInput) -> list[tuple[str, list[tuple]]]:
    try:
        workbook = load_workbook(BytesIO(item.content), read_only=True, data_only=True)
    except Exception as exc:
        raise WorkbookInspectionError("Excel 文件无法解析") from exc
    if len(workbook.sheetnames) > MAX_SHEETS:
        raise WorkbookInspectionError("Excel 工作表超过 50 个")
    output = []
    for sheet in workbook.worksheets:
        if sheet.max_column > MAX_COLUMNS or sheet.max_row > MAX_ROWS:
            raise WorkbookInspectionError("Excel 工作表超过行列限制")
        output.append((sheet.title, [tuple(row) for row in sheet.iter_rows(values_only=True)]))
    workbook.close()
    return output


def _xls_rows(item: WorkbookInput) -> list[tuple[str, list[tuple]]]:
    try:
        workbook = xlrd.open_workbook(file_contents=item.content, on_demand=True)
    except Exception as exc:
        raise WorkbookInspectionError("Excel 文件无法解析") from exc
    if workbook.nsheets > MAX_SHEETS:
        raise WorkbookInspectionError("Excel 工作表超过 50 个")
    output = []
    for sheet in workbook.sheets():
        if sheet.nrows > MAX_ROWS or sheet.ncols > MAX_COLUMNS:
            raise WorkbookInspectionError("Excel 工作表超过行列限制")
        output.append((sheet.name, [tuple(sheet.row_values(i)) for i in range(sheet.nrows)]))
    workbook.release_resources()
    return output


def _inspect_file(item: WorkbookInput) -> InspectedFile:
    _validate_input(item)
    source_sheets = _xlsx_rows(item) if item.filename.lower().endswith(".xlsx") else _xls_rows(item)
    tables: list[DetectedTable] = []
    detected_roles: set[str] = set()
    all_months: set[str] = set()
    for sheet_name, rows in source_sheets:
        found = None
        for offset, row in enumerate(rows[:20], start=1):
            columns = _map_headers(row)
            role = _role(columns)
            if role:
                found = (offset, columns, role)
                break
        if not found:
            continue
        header_row, columns, role = found
        data_rows = rows[header_row:]
        months = sorted({month for row in data_rows if (month := _month(row[columns["date"]]))})
        all_months.update(months)
        detected_roles.add(role)
        tables.append(DetectedTable(role, item.filename, sheet_name, header_row, columns, data_rows, months))
    if not tables or len(detected_roles) != 1:
        raise WorkbookInspectionError("无法从 Excel 中确定唯一的流水类型和表头")
    role = next(iter(detected_roles))
    return InspectedFile(item.filename, role, tables, sorted(all_months))


def inspect_workbooks(files: list[WorkbookInput]) -> PreflightResult:
    if len(files) != 2:
        raise WorkbookInspectionError("必须同时上传两份 Excel 文件")
    inspected = [_inspect_file(item) for item in files]
    by_role = {item.role: item for item in inspected}
    if set(by_role) != {"receipt", "expense"}:
        raise WorkbookInspectionError("必须提供一份已收流水和一份费用流水")
    receipt_months = set(by_role["receipt"].months)
    expense_months = set(by_role["expense"].months)
    return PreflightResult(
        files=inspected,
        common_months=sorted(receipt_months & expense_months),
        receipt_only_months=sorted(receipt_months - expense_months),
        expense_only_months=sorted(expense_months - receipt_months),
    )


async def inspect_workbooks_with_ai(files: list[WorkbookInput]) -> PreflightResult:
    """确定性识别优先；仅对无法识别的文件调用匿名 DeepSeek 列映射。"""
    if len(files) != 2:
        raise WorkbookInspectionError("必须同时上传两份 Excel 文件")
    inspected: list[InspectedFile] = []
    for item in files:
        try:
            inspected.append(_inspect_file(item))
            continue
        except WorkbookInspectionError as deterministic_error:
            _validate_input(item)
            source_sheets = _xlsx_rows(item) if item.filename.lower().endswith(".xlsx") else _xls_rows(item)
            from .ai_mapping import UtilityMappingError, ai_column_mapping
            try:
                mapping = await ai_column_mapping(item.filename, dict(source_sheets))
            except UtilityMappingError as exc:
                raise WorkbookInspectionError(str(exc)) from deterministic_error
            required = {"date", "receipt_amount" if mapping.role == "receipt" else "expense_amount"}
            if not required <= set(mapping.columns):
                raise WorkbookInspectionError("AI 认列缺少日期或金额列")
            rows = dict(source_sheets)[mapping.sheet]
            if mapping.header_row >= len(rows):
                raise WorkbookInspectionError("AI 返回的表头行超出工作表范围")
            data_rows = rows[mapping.header_row + 1:]
            months = sorted({month for row in data_rows if mapping.columns["date"] < len(row) and (month := _month(row[mapping.columns["date"]]))})
            table = DetectedTable(mapping.role, item.filename, mapping.sheet, mapping.header_row + 1, mapping.columns, data_rows, months)
            inspected.append(InspectedFile(item.filename, mapping.role, [table], months, "mapped_by_ai"))
    by_role = {item.role: item for item in inspected}
    if set(by_role) != {"receipt", "expense"}:
        raise WorkbookInspectionError("必须提供一份已收流水和一份费用流水")
    receipt_months = set(by_role["receipt"].months)
    expense_months = set(by_role["expense"].months)
    return PreflightResult(inspected, sorted(receipt_months & expense_months), sorted(receipt_months - expense_months), sorted(expense_months - receipt_months))
